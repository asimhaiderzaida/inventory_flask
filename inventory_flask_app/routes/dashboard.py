import logging
from flask import redirect, url_for
from flask import Blueprint, render_template, request, jsonify
from flask_login import login_required, current_user
from ..models import (
    Product, ProductInstance, db, CustomerOrderTracking,
    TenantSettings, SaleTransaction, Invoice, Customer, Return,
    CustomerOrder, UnitCost,
)
from sqlalchemy import func, text, case, literal_column
from sqlalchemy.orm import joinedload
from datetime import datetime, timedelta, timezone, date as date_type
from inventory_flask_app.utils import get_now_for_tenant
from collections import defaultdict
from inventory_flask_app.utils.mail_utils import get_low_stock_parts, get_overdue_units
from sqlalchemy.orm import aliased
import time as _time

logger = logging.getLogger(__name__)

reserved_order = aliased(CustomerOrderTracking)

dashboard_bp = Blueprint('dashboard_bp', __name__)

# Simple per-tenant in-memory cache for dashboard_stats (60-second TTL)
_stats_cache: dict = {}
_STATS_CACHE_TTL = 60


def _month_bounds(y, m):
    """Return (start datetime, end datetime) for a given year/month."""
    start = datetime(y, m, 1)
    if m == 12:
        end = datetime(y + 1, 1, 1)
    else:
        end = datetime(y, m + 1, 1)
    return start, end


def _prev_month(y, m):
    if m == 1:
        return y - 1, 12
    return y, m - 1


@dashboard_bp.route('/main_dashboard')
@login_required
def main_dashboard():
    tid = current_user.tenant_id
    today = get_now_for_tenant().date()
    now_dt = datetime.combine(today, datetime.max.time())

    # ── Month boundaries ──────────────────────────────────────────────────────
    month_start, month_end = _month_bounds(today.year, today.month)
    prev_y, prev_m = _prev_month(today.year, today.month)
    prev_month_start, prev_month_end = _month_bounds(prev_y, prev_m)

    # ── Inventory counts ──────────────────────────────────────────────────────
    # Reserved instance IDs (excluded from total inventory count)
    reserved_ids_sq = (
        db.session.query(CustomerOrderTracking.product_instance_id)
        .join(ProductInstance, CustomerOrderTracking.product_instance_id == ProductInstance.id)
        .filter(
            ProductInstance.tenant_id == tid,
            CustomerOrderTracking.status.in_(['reserved', 'delivered']),
        )
        .scalar_subquery()
    )

    reserved_count = (
        db.session.query(CustomerOrderTracking.product_instance_id)
        .join(ProductInstance, CustomerOrderTracking.product_instance_id == ProductInstance.id)
        .filter(
            ProductInstance.tenant_id == tid,
            CustomerOrderTracking.status == 'reserved',
        )
        .count()
    )

    total_inventory = (
        ProductInstance.query.join(Product)
        .filter(
            Product.tenant_id == tid,
            ProductInstance.is_sold == False,
            ~ProductInstance.id.in_(reserved_ids_sq),
        )
        .count()
    )

    under_process = (
        ProductInstance.query.join(Product)
        .outerjoin(
            reserved_order,
            (reserved_order.product_instance_id == ProductInstance.id) &
            (reserved_order.status.ilike('reserved'))
        )
        .filter(
            ProductInstance.status == 'under_process',
            ProductInstance.is_sold == False,
            reserved_order.id == None,
            Product.tenant_id == tid,
        ).count()
    )

    unprocessed = (
        ProductInstance.query.join(Product)
        .filter(
            Product.tenant_id == tid,
            ProductInstance.status == 'unprocessed',
            ProductInstance.is_sold == False,
        ).count()
    )

    # Processing backlog = unprocessed units sitting > 3 days
    three_days_ago = datetime.combine(today - timedelta(days=3), datetime.min.time())
    stale_backlog = (
        ProductInstance.query.join(Product)
        .filter(
            Product.tenant_id == tid,
            ProductInstance.status == 'unprocessed',
            ProductInstance.is_sold == False,
            ProductInstance.created_at < three_days_ago,
        ).count()
    )

    idle_units_count = (
        ProductInstance.query.join(Product)
        .filter(ProductInstance.status == 'idle', Product.tenant_id == tid)
        .count()
    )

    processed_count = (
        ProductInstance.query.join(Product)
        .filter(
            Product.tenant_id == tid,
            ProductInstance.status == 'processed',
            ProductInstance.is_sold == False,
        ).count()
    )

    disputed_count = (
        ProductInstance.query.join(Product)
        .filter(
            Product.tenant_id == tid,
            ProductInstance.status == 'disputed',
            ProductInstance.is_sold == False,
        ).count()
    )

    status_breakdown = {
        'unprocessed':  unprocessed,
        'under_process': under_process,
        'processed':    processed_count,
        'idle':         idle_units_count,
        'disputed':     disputed_count,
    }

    pending_orders = (
        CustomerOrderTracking.query
        .join(ProductInstance, CustomerOrderTracking.product_instance_id == ProductInstance.id)
        .join(Product, ProductInstance.product_id == Product.id)
        .filter(
            CustomerOrderTracking.status.in_(('reserved', 'pending')),
            Product.tenant_id == tid,
        ).count()
    )

    open_orders_count = CustomerOrder.query.filter_by(tenant_id=tid, status='open').count()

    # ── Dashboard section visibility prefs ───────────────────────────────────
    _dash_pref_rows = TenantSettings.query.filter(
        TenantSettings.tenant_id == tid,
        TenantSettings.key.in_([
            'dash_show_alert_banners', 'dash_show_key_metrics', 'dash_show_operations',
            'dash_show_sales_chart', 'dash_show_recent_sales', 'dash_show_alerts_panel',
            'dash_show_tech_workload', 'dash_show_top_models', 'dash_show_stock_aging',
            'dash_show_inventory_value',
        ])
    ).all()
    _dash_map = {r.key: r.value for r in _dash_pref_rows}
    dash_prefs = {
        'alert_banners':   _dash_map.get('dash_show_alert_banners',   'true') != 'false',
        'key_metrics':     _dash_map.get('dash_show_key_metrics',      'true') != 'false',
        'operations':      _dash_map.get('dash_show_operations',       'true') != 'false',
        'sales_chart':     _dash_map.get('dash_show_sales_chart',      'true') != 'false',
        'recent_sales':    _dash_map.get('dash_show_recent_sales',     'true') != 'false',
        'alerts_panel':    _dash_map.get('dash_show_alerts_panel',     'true') != 'false',
        'tech_workload':   _dash_map.get('dash_show_tech_workload',    'true') != 'false',
        'top_models':      _dash_map.get('dash_show_top_models',       'true') != 'false',
        'stock_aging':     _dash_map.get('dash_show_stock_aging',      'true') != 'false',
        'inventory_value': _dash_map.get('dash_show_inventory_value',  'true') != 'false',
    }

    # Aged inventory
    _aged_setting = TenantSettings.query.filter_by(tenant_id=tid, key='aged_threshold_days').first()
    aged_threshold_days = int((_aged_setting.value if _aged_setting else None) or 60)
    aged_cutoff = datetime.now(timezone.utc).replace(tzinfo=None) - timedelta(days=aged_threshold_days)
    aged_inventory_count = (
        ProductInstance.query.join(Product)
        .filter(
            Product.tenant_id == tid,
            ProductInstance.is_sold == False,
            ProductInstance.created_at < aged_cutoff,
        ).count()
    )

    # ── Stock Aging Breakdown ─────────────────────────────────────────────────
    if dash_prefs['stock_aging']:
        _now_age = datetime.now(timezone.utc).replace(tzinfo=None)
        _age7    = _now_age - timedelta(days=7)
        _age30   = _now_age - timedelta(days=30)
        _age60   = _now_age - timedelta(days=60)
        _base_age = (
            ProductInstance.query.join(Product)
            .filter(Product.tenant_id == tid, ProductInstance.is_sold == False)
        )
        aging_buckets = {
            '0_7':    _base_age.filter(ProductInstance.created_at >= _age7).count(),
            '7_30':   _base_age.filter(ProductInstance.created_at >= _age30, ProductInstance.created_at < _age7).count(),
            '30_60':  _base_age.filter(ProductInstance.created_at >= _age60, ProductInstance.created_at < _age30).count(),
            '60_plus': _base_age.filter(ProductInstance.created_at < _age60).count(),
        }
    else:
        aging_buckets = {'0_7': 0, '7_30': 0, '30_60': 0, '60_plus': 0}

    # ── Top Models ────────────────────────────────────────────────────────────
    if dash_prefs['top_models']:
        _model_label = func.coalesce(Product.model, Product.item_name)
        _top_raw = (
            db.session.query(_model_label.label('name'), func.count(ProductInstance.id).label('cnt'))
            .join(ProductInstance, ProductInstance.product_id == Product.id)
            .filter(
                Product.tenant_id == tid,
                ProductInstance.is_sold == False,
                _model_label.isnot(None),
                _model_label != 'nan',
            )
            .group_by(_model_label)
            .order_by(func.count(ProductInstance.id).desc())
            .limit(6)
            .all()
        )
        top_models = [
            {'name': r.name, 'count': r.cnt}
            for r in _top_raw
            if r.name and r.name.lower() != 'nan'
        ][:5]
    else:
        top_models = []

    # ── Processing Throughput ─────────────────────────────────────────────────
    from inventory_flask_app.models import ProductProcessLog as _PPLog
    _week_start      = datetime.combine(today - timedelta(days=today.weekday()), datetime.min.time())
    _last_week_start = _week_start - timedelta(days=7)

    throughput_this_week = (
        db.session.query(func.count(_PPLog.id))
        .join(ProductInstance, _PPLog.product_instance_id == ProductInstance.id)
        .join(Product, ProductInstance.product_id == Product.id)
        .filter(
            Product.tenant_id == tid,
            _PPLog.to_stage == 'processed',
            _PPLog.moved_at >= _week_start,
        )
        .scalar() or 0
    )
    throughput_last_week = (
        db.session.query(func.count(_PPLog.id))
        .join(ProductInstance, _PPLog.product_instance_id == ProductInstance.id)
        .join(Product, ProductInstance.product_id == Product.id)
        .filter(
            Product.tenant_id == tid,
            _PPLog.to_stage == 'processed',
            _PPLog.moved_at >= _last_week_start,
            _PPLog.moved_at < _week_start,
        )
        .scalar() or 0
    )
    _spark_start = datetime.combine(today - timedelta(days=13), datetime.min.time())
    _spark_rows = (
        db.session.query(
            func.date(_PPLog.moved_at).label('day'),
            func.count(_PPLog.id).label('cnt'),
        )
        .join(ProductInstance, _PPLog.product_instance_id == ProductInstance.id)
        .join(Product, ProductInstance.product_id == Product.id)
        .filter(
            Product.tenant_id == tid,
            _PPLog.to_stage == 'processed',
            _PPLog.moved_at >= _spark_start,
        )
        .group_by(func.date(_PPLog.moved_at))
        .all()
    )
    _spark_map  = {r.day: r.cnt for r in _spark_rows}
    _spark_days = [today - timedelta(days=i) for i in range(13, -1, -1)]
    throughput_sparkline = [_spark_map.get(d, 0) for d in _spark_days]

    # ── Inventory Value ───────────────────────────────────────────────────────
    if dash_prefs['inventory_value']:
        _inv_val = (
            db.session.query(
                func.coalesce(func.sum(UnitCost.total_cost), 0).label('cost'),
                func.coalesce(func.sum(ProductInstance.asking_price), 0).label('asking'),
            )
            .join(ProductInstance, UnitCost.instance_id == ProductInstance.id)
            .join(Product, ProductInstance.product_id == Product.id)
            .filter(Product.tenant_id == tid, ProductInstance.is_sold == False)
            .first()
        )
        inventory_cost   = round(float(_inv_val.cost   or 0), 2)
        inventory_asking = round(float(_inv_val.asking or 0), 2)
        priced_count = (
            ProductInstance.query.join(Product)
            .filter(
                Product.tenant_id == tid,
                ProductInstance.is_sold == False,
                ProductInstance.asking_price > 0,
            )
            .count()
        )
    else:
        inventory_cost = inventory_asking = 0.0
        priced_count = 0

    # ── Sales / Revenue KPIs ──────────────────────────────────────────────────
    def _revenue_query(start, end):
        return db.session.query(
            func.coalesce(func.sum(SaleTransaction.price_at_sale), 0)
        ).join(ProductInstance, SaleTransaction.product_instance_id == ProductInstance.id
        ).join(Product, ProductInstance.product_id == Product.id
        ).filter(
            Product.tenant_id == tid,
            SaleTransaction.date_sold >= start,
            SaleTransaction.date_sold < end,
        ).scalar()

    def _units_sold_query(start, end):
        return (
            SaleTransaction.query
            .join(ProductInstance, SaleTransaction.product_instance_id == ProductInstance.id)
            .join(Product, ProductInstance.product_id == Product.id)
            .filter(
                Product.tenant_id == tid,
                SaleTransaction.date_sold >= start,
                SaleTransaction.date_sold < end,
            ).count()
        )

    revenue_this_month   = round(float(_revenue_query(month_start, month_end)), 2)
    revenue_last_month   = round(float(_revenue_query(prev_month_start, prev_month_end)), 2)
    units_sold_this_month = _units_sold_query(month_start, month_end)
    units_sold_last_month = _units_sold_query(prev_month_start, prev_month_end)

    # Trend % (None if no prior data)
    def _trend(current, previous):
        if previous == 0:
            return None
        return round((current - previous) / previous * 100, 1)

    revenue_trend = _trend(revenue_this_month, revenue_last_month)
    sales_trend   = _trend(units_sold_this_month, units_sold_last_month)

    # ── Returns count ─────────────────────────────────────────────────────────
    returns_this_month = Return.query.filter(
        Return.tenant_id == tid,
        Return.return_date >= datetime.combine(month_start, datetime.min.time()),
    ).count()

    # ── Accounting KPIs ───────────────────────────────────────────────────────
    from inventory_flask_app.models import AccountReceivable, Expense, OtherIncome

    outstanding_ar = round(float(
        db.session.query(
            func.coalesce(func.sum(AccountReceivable.amount_due - AccountReceivable.amount_paid), 0)
        ).filter(
            AccountReceivable.tenant_id == tid,
            AccountReceivable.status.in_(('open', 'partial', 'overdue')),
        ).scalar() or 0
    ), 2)

    _month_start_date = today.replace(day=1)
    mtd_expenses = float(
        db.session.query(func.coalesce(func.sum(Expense.amount), 0))
        .filter(Expense.tenant_id == tid, Expense.deleted_at.is_(None),
                Expense.expense_date >= _month_start_date).scalar() or 0
    )
    mtd_other_income = float(
        db.session.query(func.coalesce(func.sum(OtherIncome.amount), 0))
        .filter(OtherIncome.tenant_id == tid, OtherIncome.deleted_at.is_(None),
                OtherIncome.income_date >= _month_start_date).scalar() or 0
    )
    net_profit_month = round(revenue_this_month + mtd_other_income - mtd_expenses, 2)

    # ── SLA / Overdue ─────────────────────────────────────────────────────────
    low_stock_parts = get_low_stock_parts(tid)
    overdue_units   = get_overdue_units(tid)
    overdue_count   = len(overdue_units)

    # ── Recently returned units (last 7 days) ─────────────────────────────────
    seven_days_ago = datetime.combine(today - timedelta(days=7), datetime.min.time())
    recently_returned_count = (
        ProductInstance.query.join(Product)
        .filter(
            Product.tenant_id == tid,
            ProductInstance.returned_at.isnot(None),
            ProductInstance.returned_at >= seven_days_ago,
            ProductInstance.is_sold == False,
        )
        .count()
    )

    # ── Recent Sales (last 5) ─────────────────────────────────────────────────
    if dash_prefs['recent_sales']:
        recent_sale_rows = (
            db.session.query(SaleTransaction, Customer, Invoice)
            .join(ProductInstance, SaleTransaction.product_instance_id == ProductInstance.id)
            .join(Product, ProductInstance.product_id == Product.id)
            .join(Customer, SaleTransaction.customer_id == Customer.id)
            .outerjoin(Invoice, SaleTransaction.invoice_id == Invoice.id)
            .filter(Product.tenant_id == tid)
            .order_by(SaleTransaction.date_sold.desc())
            .limit(5)
            .all()
        )
        recent_sales = []
        for txn, cust, inv in recent_sale_rows:
            recent_sales.append({
                'date':           txn.date_sold,
                'customer':       cust.name if cust else '—',
                'amount':         float(txn.price_at_sale or 0),
                'payment_method': txn.payment_method or '—',
                'invoice_id':     inv.id if inv else None,
                'invoice_number': inv.invoice_number if inv else None,
            })
    else:
        recent_sales = []

    # ── Technician Workload ───────────────────────────────────────────────────
    from inventory_flask_app.models import User as UserModel, ProductProcessLog
    if not dash_prefs['tech_workload']:
        tech_workload = []
    else:
        tech_rows = (
            db.session.query(
                UserModel.id,
                UserModel.username,
                UserModel.full_name,
                func.count(ProductInstance.id).label('unit_count'),
                func.min(ProductInstance.entered_stage_at).label('oldest_stage_at'),
                func.min(ProductInstance.created_at).label('oldest_created'),
            )
            .join(ProductInstance, ProductInstance.assigned_to_user_id == UserModel.id)
            .join(Product, ProductInstance.product_id == Product.id)
            .filter(
                Product.tenant_id == tid,
                ProductInstance.status.in_(['under_process']),
                ProductInstance.is_sold == False,
            )
            .group_by(UserModel.id, UserModel.username, UserModel.full_name)
            .order_by(func.count(ProductInstance.id).desc())
            .limit(10)
            .all()
        )
        avg_times = {}
        if tech_rows:
            tech_ids = [r.id for r in tech_rows]
            avg_rows = (
                db.session.query(
                    ProductProcessLog.moved_by,
                    func.avg(ProductProcessLog.duration_minutes).label('avg_min'),
                )
                .filter(
                    ProductProcessLog.moved_by.in_(tech_ids),
                    ProductProcessLog.duration_minutes.isnot(None),
                )
                .group_by(ProductProcessLog.moved_by)
                .all()
            )
            avg_times = {r.moved_by: round(float(r.avg_min), 0) for r in avg_rows}
        tech_workload = []
        for r in tech_rows:
            oldest_dt = r.oldest_stage_at or r.oldest_created
            if oldest_dt:
                age_hours = int((datetime.now(timezone.utc).replace(tzinfo=None) - oldest_dt).total_seconds() / 3600)
            else:
                age_hours = None
            avg_min = avg_times.get(r.id)
            tech_workload.append({
                'name':        r.full_name or r.username,
                'count':       r.unit_count,
                'age_hours':   age_hours,
                'avg_minutes': int(avg_min) if avg_min else None,
            })

    # ── Technician: my assigned units + today's completions ──────────────────
    from inventory_flask_app.models import ProductProcessLog, ProcessStage
    my_units = []
    my_completions_today = 0
    if current_user.role == 'technician':
        today_start = datetime.combine(today, datetime.min.time())
        my_units_raw = (
            ProductInstance.query.join(Product)
            .filter(
                Product.tenant_id == tid,
                ProductInstance.assigned_to_user_id == current_user.id,
                ProductInstance.status == 'under_process',
                ProductInstance.is_sold == False,
            )
            .order_by(ProductInstance.entered_stage_at.asc().nullslast())
            .limit(30)
            .all()
        )
        # Build SLA info for each unit
        stages = ProcessStage.query.filter_by(tenant_id=tid).all()
        sla_map = {s.name: s.sla_hours for s in stages if s.sla_hours and s.sla_hours > 0}
        now_utc = datetime.now(timezone.utc).replace(tzinfo=None)
        for inst in my_units_raw:
            sla_h = sla_map.get((inst.process_stage or '').strip(), 0)
            hours_in = None
            is_overdue = False
            if inst.entered_stage_at:
                since = inst.entered_stage_at
                if hasattr(since, 'utcoffset') and since.utcoffset() is not None:
                    since = since.replace(tzinfo=None) - since.utcoffset()
                hours_in = round((now_utc - since).total_seconds() / 3600, 1)
                is_overdue = sla_h > 0 and hours_in > sla_h
            my_units.append({
                'instance': inst,
                'serial': inst.serial,
                'stage': inst.process_stage or '—',
                'sla_hours': sla_h,
                'hours_in': hours_in,
                'is_overdue': is_overdue,
            })
        my_completions_today = (
            ProductProcessLog.query
            .filter(
                ProductProcessLog.moved_by == current_user.id,
                ProductProcessLog.moved_at >= today_start,
            )
            .count()
        )

    # ── Chart data (30-day default, JS handles 7d/90d slicing) ───────────────
    if dash_prefs['sales_chart']:
        chart_days = 90  # send 90 days; JS will slice to 7/30/90
        chart_start = datetime.combine(today - timedelta(days=chart_days - 1), datetime.min.time())
        chart_end   = datetime.combine(today + timedelta(days=1), datetime.min.time())
        daily_rows = (
            db.session.query(
                func.date(SaleTransaction.date_sold).label('sale_day'),
                func.count(SaleTransaction.id).label('cnt'),
                func.sum(SaleTransaction.price_at_sale).label('revenue'),
            )
            .join(ProductInstance, SaleTransaction.product_instance_id == ProductInstance.id)
            .join(Product, ProductInstance.product_id == Product.id)
            .filter(
                Product.tenant_id == tid,
                SaleTransaction.date_sold >= chart_start,
                SaleTransaction.date_sold < chart_end,
            )
            .group_by(func.date(SaleTransaction.date_sold))
            .all()
        )
        cnt_map = {r.sale_day: r.cnt for r in daily_rows}
        rev_map = {r.sale_day: round(float(r.revenue or 0), 2) for r in daily_rows}
        all_days = [today - timedelta(days=i) for i in range(chart_days - 1, -1, -1)]
        sales_chart = {
            'labels':  [d.strftime('%Y-%m-%d') for d in all_days],
            'sales':   [cnt_map.get(d, 0) for d in all_days],
            'revenue': [rev_map.get(d, 0) for d in all_days],
        }
    else:
        sales_chart = {'labels': [], 'sales': [], 'revenue': []}

    return render_template('main_dashboard.html',
        # Primary KPIs
        total_inventory=total_inventory,
        reserved_count=reserved_count,
        revenue_this_month=revenue_this_month,
        revenue_last_month=revenue_last_month,
        revenue_trend=revenue_trend,
        units_sold_this_month=units_sold_this_month,
        units_sold_last_month=units_sold_last_month,
        sales_trend=sales_trend,
        net_profit_month=net_profit_month,
        # Operations
        under_process=under_process,
        unprocessed=unprocessed,
        stale_backlog=stale_backlog,
        outstanding_ar=outstanding_ar,
        overdue_count=overdue_count,
        overdue_units=overdue_units,
        returns_this_month=returns_this_month,
        # Alerts
        low_stock_parts=low_stock_parts,
        recently_returned_count=recently_returned_count,
        aged_inventory_count=aged_inventory_count,
        idle_units_count=idle_units_count,
        pending_orders=pending_orders,
        open_orders_count=open_orders_count,
        # Tables
        recent_sales=recent_sales,
        tech_workload=tech_workload,
        # Chart
        sales_chart=sales_chart,
        # Technician role
        my_units=my_units,
        my_completions_today=my_completions_today,
        # Configurable thresholds
        aged_threshold_days=aged_threshold_days,
        # New: inventory breakdown
        status_breakdown=status_breakdown,
        processed_count=processed_count,
        disputed_count=disputed_count,
        # New: aging
        aging_buckets=aging_buckets,
        # New: top models
        top_models=top_models,
        # New: throughput
        throughput_this_week=throughput_this_week,
        throughput_last_week=throughput_last_week,
        throughput_sparkline=throughput_sparkline,
        # New: inventory value
        inventory_cost=inventory_cost,
        inventory_asking=inventory_asking,
        priced_count=priced_count,
    )


@dashboard_bp.route('/')
def home_redirect():
    return redirect(url_for('dashboard_bp.main_dashboard'))


def _get_export_data(tid):
    """Gather KPI snapshot data for Excel/PDF export."""
    today = get_now_for_tenant().date()
    month_start, month_end = _month_bounds(today.year, today.month)
    prev_y, prev_m = _prev_month(today.year, today.month)
    prev_month_start, prev_month_end = _month_bounds(prev_y, prev_m)

    from inventory_flask_app.models import (
        AccountReceivable, Expense, OtherIncome,
        ProductProcessLog, User as UserModel,
    )

    # Inventory status counts
    status_counts = dict(
        db.session.query(ProductInstance.status, func.count(ProductInstance.id))
        .join(Product)
        .filter(Product.tenant_id == tid, ProductInstance.is_sold == False)
        .group_by(ProductInstance.status)
        .all()
    )

    total_inventory = sum(status_counts.values())

    # Revenue / sales
    def _rev(start, end):
        return round(float(
            db.session.query(func.coalesce(func.sum(SaleTransaction.price_at_sale), 0))
            .join(ProductInstance, SaleTransaction.product_instance_id == ProductInstance.id)
            .join(Product, ProductInstance.product_id == Product.id)
            .filter(Product.tenant_id == tid,
                    SaleTransaction.date_sold >= start,
                    SaleTransaction.date_sold < end).scalar() or 0
        ), 2)

    def _units(start, end):
        return (
            SaleTransaction.query
            .join(ProductInstance, SaleTransaction.product_instance_id == ProductInstance.id)
            .join(Product, ProductInstance.product_id == Product.id)
            .filter(Product.tenant_id == tid,
                    SaleTransaction.date_sold >= start,
                    SaleTransaction.date_sold < end).count()
        )

    revenue_mtd = _rev(month_start, month_end)
    revenue_prev = _rev(prev_month_start, prev_month_end)
    units_mtd = _units(month_start, month_end)

    # Expenses / other income
    month_start_date = today.replace(day=1)
    mtd_expenses = float(
        db.session.query(func.coalesce(func.sum(Expense.amount), 0))
        .filter(Expense.tenant_id == tid, Expense.deleted_at.is_(None),
                Expense.expense_date >= month_start_date).scalar() or 0
    )
    mtd_other_income = float(
        db.session.query(func.coalesce(func.sum(OtherIncome.amount), 0))
        .filter(OtherIncome.tenant_id == tid, OtherIncome.deleted_at.is_(None),
                OtherIncome.income_date >= month_start_date).scalar() or 0
    )
    net_profit = round(revenue_mtd + mtd_other_income - mtd_expenses, 2)

    # AR
    outstanding_ar = round(float(
        db.session.query(
            func.coalesce(func.sum(AccountReceivable.amount_due - AccountReceivable.amount_paid), 0)
        ).filter(AccountReceivable.tenant_id == tid,
                 AccountReceivable.status.in_(('open', 'partial', 'overdue'))).scalar() or 0
    ), 2)

    # Overdue
    overdue_units = get_overdue_units(tid)
    overdue_count = len(overdue_units)

    # Returns this month
    returns_mtd = Return.query.filter(
        Return.tenant_id == tid,
        Return.return_date >= datetime.combine(month_start, datetime.min.time()),
    ).count()

    # Aging buckets
    _now = datetime.now(timezone.utc).replace(tzinfo=None)
    _base = ProductInstance.query.join(Product).filter(
        Product.tenant_id == tid, ProductInstance.is_sold == False
    )
    aging_buckets = {
        '0–7 days':   _base.filter(ProductInstance.created_at >= _now - timedelta(days=7)).count(),
        '7–30 days':  _base.filter(ProductInstance.created_at >= _now - timedelta(days=30),
                                    ProductInstance.created_at < _now - timedelta(days=7)).count(),
        '30–60 days': _base.filter(ProductInstance.created_at >= _now - timedelta(days=60),
                                    ProductInstance.created_at < _now - timedelta(days=30)).count(),
        '60+ days':   _base.filter(ProductInstance.created_at < _now - timedelta(days=60)).count(),
    }

    # Inventory value
    inv_val = (
        db.session.query(
            func.coalesce(func.sum(UnitCost.total_cost), 0).label('cost'),
            func.coalesce(func.sum(ProductInstance.asking_price), 0).label('asking'),
        )
        .join(ProductInstance, UnitCost.instance_id == ProductInstance.id)
        .join(Product, ProductInstance.product_id == Product.id)
        .filter(Product.tenant_id == tid, ProductInstance.is_sold == False)
        .first()
    )
    inventory_cost   = round(float(inv_val.cost   or 0), 2)
    inventory_asking = round(float(inv_val.asking or 0), 2)

    # Recent sales (last 10)
    recent_sale_rows = (
        db.session.query(SaleTransaction, Customer, Invoice)
        .join(ProductInstance, SaleTransaction.product_instance_id == ProductInstance.id)
        .join(Product, ProductInstance.product_id == Product.id)
        .join(Customer, SaleTransaction.customer_id == Customer.id)
        .outerjoin(Invoice, SaleTransaction.invoice_id == Invoice.id)
        .filter(Product.tenant_id == tid)
        .order_by(SaleTransaction.date_sold.desc())
        .limit(10)
        .all()
    )
    recent_sales = [
        {
            'date':           txn.date_sold.strftime('%d %b %Y') if txn.date_sold else '—',
            'customer':       cust.name if cust else '—',
            'amount':         float(txn.price_at_sale or 0),
            'payment_method': txn.payment_method or '—',
            'invoice_number': inv.invoice_number if inv else '—',
        }
        for txn, cust, inv in recent_sale_rows
    ]

    # Tech workload
    tech_rows = (
        db.session.query(
            UserModel.id,
            UserModel.username,
            UserModel.full_name,
            func.count(ProductInstance.id).label('unit_count'),
        )
        .join(ProductInstance, ProductInstance.assigned_to_user_id == UserModel.id)
        .join(Product, ProductInstance.product_id == Product.id)
        .filter(
            Product.tenant_id == tid,
            ProductInstance.status == 'under_process',
            ProductInstance.is_sold == False,
        )
        .group_by(UserModel.id, UserModel.username, UserModel.full_name)
        .order_by(func.count(ProductInstance.id).desc())
        .limit(15)
        .all()
    )
    avg_time_rows = {}
    if tech_rows:
        avg_time_rows = dict(
            db.session.query(
                ProductProcessLog.moved_by,
                func.avg(ProductProcessLog.duration_minutes).label('avg_min'),
            )
            .filter(
                ProductProcessLog.moved_by.in_([r.id for r in tech_rows]),
                ProductProcessLog.duration_minutes.isnot(None),
            )
            .group_by(ProductProcessLog.moved_by)
            .all()
        )
    tech_workload = [
        {
            'name':       r.full_name or r.username,
            'count':      r.unit_count,
            'avg_minutes': round(float(avg_time_rows[r.id]), 0) if r.id in avg_time_rows else None,
        }
        for r in tech_rows
    ]

    # Tenant settings for display
    _settings = {
        s.key: s.value
        for s in TenantSettings.query.filter_by(tenant_id=tid).all()
    }

    return {
        'generated_at':    today.strftime('%d %B %Y'),
        'month_label':     today.strftime('%B %Y'),
        'company_name':    _settings.get('company_name', ''),
        'currency':        _settings.get('currency', 'AED'),
        # KPIs
        'total_inventory': total_inventory,
        'revenue_mtd':     revenue_mtd,
        'revenue_prev':    revenue_prev,
        'units_mtd':       units_mtd,
        'net_profit':      net_profit,
        'outstanding_ar':  outstanding_ar,
        'overdue_count':   overdue_count,
        'returns_mtd':     returns_mtd,
        'inventory_cost':  inventory_cost,
        'inventory_asking': inventory_asking,
        # Breakdowns
        'status_counts':   status_counts,
        'aging_buckets':   aging_buckets,
        'recent_sales':    recent_sales,
        'tech_workload':   tech_workload,
    }


@dashboard_bp.route('/dashboard/export/excel')
@login_required
def dashboard_export_excel():
    if current_user.role not in ('admin', 'supervisor'):
        from flask import abort
        abort(403)

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        from flask import flash
        flash('openpyxl is not installed. Run: pip install openpyxl', 'danger')
        return redirect(url_for('dashboard_bp.main_dashboard'))

    from flask import make_response
    import io

    tid = current_user.tenant_id
    d = _get_export_data(tid)
    currency = d['currency']
    wb = openpyxl.Workbook()

    # ── Styles ────────────────────────────────────────────────────────────────
    hdr_fill  = PatternFill('solid', fgColor='1D4ED8')
    hdr_font  = Font(bold=True, color='FFFFFF', size=11)
    bold_font = Font(bold=True, size=11)
    h1_font   = Font(bold=True, size=14)
    muted     = Font(color='6B7280', size=10)
    cent      = Alignment(horizontal='center')
    thin      = Side(style='thin', color='D1D5DB')
    thin_bdr  = Border(bottom=Side(style='thin', color='E5E7EB'))

    def _header_row(ws, headers, row=1):
        for col, text in enumerate(headers, 1):
            c = ws.cell(row=row, column=col, value=text)
            c.font, c.fill, c.alignment = hdr_font, hdr_fill, cent

    def _autowidth(ws):
        for col in ws.columns:
            max_len = max((len(str(cell.value or '')) for cell in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

    # ── Sheet 1: KPI Summary ──────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'KPI Summary'
    ws1['A1'] = d['company_name'] or 'Dashboard Export'
    ws1['A1'].font = h1_font
    ws1['A2'] = f"Generated: {d['generated_at']}  |  Period: {d['month_label']}"
    ws1['A2'].font = muted
    ws1.append([])

    kpi_rows = [
        ('Total Inventory (units)',  d['total_inventory'],  ''),
        (f'MTD Revenue ({currency})',  d['revenue_mtd'],  f"prev month: {currency} {d['revenue_prev']:,.2f}"),
        ('MTD Units Sold',           d['units_mtd'],      ''),
        (f'Net Profit MTD ({currency})', d['net_profit'], ''),
        (f'Outstanding AR ({currency})', d['outstanding_ar'], ''),
        ('Overdue Units',            d['overdue_count'],  ''),
        ('Returns This Month',       d['returns_mtd'],    ''),
        (f'Inventory Cost ({currency})',   d['inventory_cost'],   ''),
        (f'Inventory Asking ({currency})', d['inventory_asking'], ''),
    ]
    _header_row(ws1, ['KPI', 'Value', 'Note'], row=4)
    for i, (label, val, note) in enumerate(kpi_rows, 5):
        ws1.cell(i, 1, label)
        ws1.cell(i, 2, val)
        ws1.cell(i, 3, note)
    _autowidth(ws1)

    # ── Sheet 2: Inventory Status ─────────────────────────────────────────────
    ws2 = wb.create_sheet('Inventory Status')
    _header_row(ws2, ['Status', 'Count'])
    for i, (status, cnt) in enumerate(d['status_counts'].items(), 2):
        ws2.cell(i, 1, status.replace('_', ' ').title())
        ws2.cell(i, 2, cnt)
    ws2.append([])
    ws2.append(['Stock Aging', ''])
    ws2[ws2.max_row][0].font = bold_font
    for label, cnt in d['aging_buckets'].items():
        ws2.append([label, cnt])
    _autowidth(ws2)

    # ── Sheet 3: Recent Sales ─────────────────────────────────────────────────
    ws3 = wb.create_sheet('Recent Sales')
    _header_row(ws3, ['Date', 'Customer', f'Amount ({currency})', 'Payment', 'Invoice #'])
    for i, row in enumerate(d['recent_sales'], 2):
        ws3.cell(i, 1, row['date'])
        ws3.cell(i, 2, row['customer'])
        ws3.cell(i, 3, row['amount'])
        ws3.cell(i, 4, row['payment_method'])
        ws3.cell(i, 5, row['invoice_number'])
    _autowidth(ws3)

    # ── Sheet 4: Tech Workload ────────────────────────────────────────────────
    ws4 = wb.create_sheet('Tech Workload')
    _header_row(ws4, ['Technician', 'Units In Progress', 'Avg Processing Time (min)'])
    for i, row in enumerate(d['tech_workload'], 2):
        ws4.cell(i, 1, row['name'])
        ws4.cell(i, 2, row['count'])
        ws4.cell(i, 3, row['avg_minutes'] or '—')
    _autowidth(ws4)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"dashboard_{d['generated_at'].replace(' ', '_')}.xlsx"
    resp = make_response(buf.read())
    resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    resp.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


@dashboard_bp.route('/dashboard/export/pdf')
@login_required
def dashboard_export_pdf():
    if current_user.role not in ('admin', 'supervisor'):
        from flask import abort
        abort(403)

    try:
        from weasyprint import HTML as WeasyprintHTML
    except ImportError:
        from flask import flash
        flash('weasyprint is not installed. Run: pip install weasyprint', 'danger')
        return redirect(url_for('dashboard_bp.main_dashboard'))

    from flask import make_response, render_template as _rt
    import io

    tid = current_user.tenant_id
    d = _get_export_data(tid)

    html_str = _rt('reports/dashboard_pdf.html', **d)
    pdf_bytes = WeasyprintHTML(string=html_str).write_pdf()

    filename = f"dashboard_{d['generated_at'].replace(' ', '_')}.pdf"
    resp = make_response(pdf_bytes)
    resp.headers['Content-Type'] = 'application/pdf'
    resp.headers['Content-Disposition'] = f'inline; filename="{filename}"'
    return resp


# ── Live stats API (60s cached) ───────────────────────────────────────────────
@dashboard_bp.route('/api/dashboard_stats')
@login_required
def dashboard_stats():
    tid = current_user.tenant_id

    cached = _stats_cache.get(tid)
    if cached and (_time.time() - cached[0]) < _STATS_CACHE_TTL:
        return jsonify(cached[1])

    today = get_now_for_tenant().date()
    month_start = datetime(today.year, today.month, 1)

    under_process = (
        ProductInstance.query.join(Product)
        .filter(Product.tenant_id == tid,
                ProductInstance.status == 'under_process',
                ProductInstance.is_sold == False).count()
    )
    unprocessed = (
        ProductInstance.query.join(Product)
        .filter(Product.tenant_id == tid,
                ProductInstance.status == 'unprocessed',
                ProductInstance.is_sold == False).count()
    )
    reserved_ids_live = (
        db.session.query(CustomerOrderTracking.product_instance_id)
        .join(ProductInstance, CustomerOrderTracking.product_instance_id == ProductInstance.id)
        .filter(
            ProductInstance.tenant_id == tid,
            CustomerOrderTracking.status.in_(['reserved', 'delivered']),
        )
        .scalar_subquery()
    )
    reserved_count = (
        db.session.query(CustomerOrderTracking.product_instance_id)
        .join(ProductInstance, CustomerOrderTracking.product_instance_id == ProductInstance.id)
        .filter(
            ProductInstance.tenant_id == tid,
            CustomerOrderTracking.status == 'reserved',
        )
        .count()
    )
    total_inventory = (
        ProductInstance.query.join(Product)
        .filter(
            Product.tenant_id == tid,
            ProductInstance.is_sold == False,
            ~ProductInstance.id.in_(reserved_ids_live),
        )
        .count()
    )
    units_sold_this_month = (
        SaleTransaction.query
        .join(ProductInstance, SaleTransaction.product_instance_id == ProductInstance.id)
        .join(Product, ProductInstance.product_id == Product.id)
        .filter(Product.tenant_id == tid,
                SaleTransaction.date_sold >= month_start).count()
    )
    revenue_this_month = round(float(
        db.session.query(func.coalesce(func.sum(SaleTransaction.price_at_sale), 0))
        .join(ProductInstance, SaleTransaction.product_instance_id == ProductInstance.id)
        .join(Product, ProductInstance.product_id == Product.id)
        .filter(Product.tenant_id == tid,
                SaleTransaction.date_sold >= month_start).scalar() or 0
    ), 2)

    overdue_count   = len(get_overdue_units(tid))
    low_stock_count = len(get_low_stock_parts(tid))

    three_days_ago = datetime.combine(today - timedelta(days=3), datetime.min.time())
    stale_backlog = (
        ProductInstance.query.join(Product)
        .filter(Product.tenant_id == tid,
                ProductInstance.status == 'unprocessed',
                ProductInstance.is_sold == False,
                ProductInstance.created_at < three_days_ago).count()
    )

    from inventory_flask_app.models import AccountReceivable
    outstanding_ar = round(float(
        db.session.query(
            func.coalesce(func.sum(AccountReceivable.amount_due - AccountReceivable.amount_paid), 0)
        ).filter(AccountReceivable.tenant_id == tid,
                 AccountReceivable.status.in_(('open', 'partial', 'overdue'))).scalar() or 0
    ), 2)

    returns_this_month = Return.query.filter(
        Return.tenant_id == tid,
        Return.return_date >= month_start,
    ).count()

    result = {
        'under_process':        under_process,
        'unprocessed':          unprocessed,
        'total_inventory':      total_inventory,
        'reserved_count':       reserved_count,
        'units_sold_this_month': units_sold_this_month,
        'revenue_this_month':   revenue_this_month,
        'overdue_count':        overdue_count,
        'low_stock_count':      low_stock_count,
        'stale_backlog':        stale_backlog,
        'outstanding_ar':       outstanding_ar,
        'returns_this_month':   returns_this_month,
    }
    _stats_cache[tid] = (_time.time(), result)
    return jsonify(result)
