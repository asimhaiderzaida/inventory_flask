
import logging
from flask import Blueprint, render_template, request, redirect, url_for, flash, session, make_response

logger = logging.getLogger(__name__)
from flask_login import login_required, current_user
from inventory_flask_app import csrf
from inventory_flask_app.models import db, Product, ProductInstance, PurchaseOrder, Vendor, Location, Bin, PartStock, Part
from inventory_flask_app.models import CustomerOrderTracking, ProductProcessLog
from flask import jsonify
from sqlalchemy.orm import aliased, joinedload
from sqlalchemy import func, or_
from datetime import datetime, timedelta
import qrcode
from io import BytesIO, StringIO
import base64
from inventory_flask_app.utils.utils import get_instance_id, calc_duration_minutes, create_notification, sync_reservation_stage, admin_required, admin_or_supervisor_required
from inventory_flask_app.utils import get_now_for_tenant
import csv

stock_bp = Blueprint('stock_bp', __name__, url_prefix='/stock')

@csrf.exempt
@stock_bp.route('/checkout/scan_add', methods=['POST'])
@login_required
def add_checkout_scan():
    from flask import jsonify
    data = request.get_json()
    if not data or not data.get("serial"):
        return jsonify({"error": "Missing serial"}), 400

    scanned = session.get('scanned_checkout', [])
    serials = [s.get('serial') for s in scanned]

    # Query instance to enrich the entry with asset, model, item_name
    instance = ProductInstance.query.join(Product).filter(
        (ProductInstance.serial == data["serial"]) | (ProductInstance.asset == data["serial"]),
        Product.tenant_id == current_user.tenant_id
    ).first()

    if not instance:
        return jsonify({"error": "Serial not found"}), 404

    new_entry = {
        "serial": data["serial"],
        "note": data.get("note", ""),
        "asset": instance.asset or "",
        "model": instance.product.model if instance.product else "",
        "item_name": instance.product.item_name if instance.product else "",
        "instance_id": instance.id
    }

    # Replace existing note if serial already present
    if data["serial"] in serials:
        for entry in scanned:
            if entry["serial"] == data["serial"]:
                entry["note"] = new_entry["note"]
                entry["asset"] = new_entry["asset"]
                entry["model"] = new_entry["model"]
                entry["item_name"] = new_entry["item_name"]
                entry["instance_id"] = new_entry["instance_id"]
    else:
        scanned.append(new_entry)

    # Always lookup the final entry (enriched) for the response
    final_entry = next((entry for entry in scanned if entry['serial'] == data['serial']), new_entry)
    if not final_entry:
        final_entry = new_entry
    session['scanned_checkout'] = scanned
    session.modified = True
    return jsonify({"message": "Scanned unit added", "count": len(scanned), "entry": final_entry})



# --- AJAX-based Check-in Scan Add Route ---
@csrf.exempt
@stock_bp.route('/checkin/scan_add', methods=['POST'])
@login_required
def add_checkin_scan():
    data = request.get_json()
    if not data or not data.get("serial"):
        return jsonify({"error": "Missing serial"}), 400

    serial = data["serial"].strip()
    scanned = session.get('scanned_checkin', [])

    # Avoid duplicate serials in session
    if any(entry["serial"] == serial for entry in scanned):
        return jsonify({
            "message": "Already scanned",
            "entry": next(e for e in scanned if e["serial"] == serial)
        })

    instance = ProductInstance.query.join(Product).filter(
        (ProductInstance.serial == serial) | (ProductInstance.asset == serial),
        Product.tenant_id == current_user.tenant_id
    ).first()

    if not instance:
        return jsonify({"error": "Serial not found"}), 404

    # fix 12: detect assignment conflict so UI can warn immediately
    conflict_user = None
    if instance.assigned_to_user_id and instance.assigned_to_user_id != current_user.id:
        from inventory_flask_app.models import User as UserModel
        assigned = db.session.get(UserModel, instance.assigned_to_user_id)
        conflict_user = assigned.username if assigned else 'another technician'

    new_entry = {
        "serial": serial,
        "note": data.get("note", ""),
        "asset": instance.asset or "",
        "model": instance.product.model if instance.product else "",
        "item_name": instance.product.item_name if instance.product else "",
        "stage": data.get("stage", ""),
        "team": data.get("team", ""),
        "instance_id": instance.id,
        "conflict_user": conflict_user,
    }

    scanned.append(new_entry)
    session['scanned_checkin'] = scanned
    session.modified = True

    return jsonify({
        "message": "Scanned unit added",
        "entry": new_entry,
        "count": len(scanned)
    })


# --- Remove from check-in scan queue ---
@csrf.exempt
@stock_bp.route('/remove_checkin_scan/<path:serial>', methods=['POST'])
@login_required
def remove_checkin_scan(serial):
    scanned = session.get('scanned_checkin', [])
    session['scanned_checkin'] = [e for e in scanned if e.get('serial') != serial]
    session.modified = True
    return redirect(url_for('stock_bp.process_stage_update', tab='check_in'))


# --- Remove from check-out scan queue ---
@csrf.exempt
@stock_bp.route('/remove_checkout_scan/<path:serial>', methods=['POST'])
@login_required
def remove_checkout_scan(serial):
    scanned = session.get('scanned_checkout', [])
    session['scanned_checkout'] = [e for e in scanned if e.get('serial') != serial]
    session.modified = True
    return redirect(url_for('stock_bp.process_stage_update', tab='check_out'))


# Pipeline routes moved to routes/pipeline.py (pipeline_bp)

# ─────────────────────────────────────────────────────────────
# Bulk Price Editor
# ─────────────────────────────────────────────────────────────
@stock_bp.route('/bulk_price')
@login_required
def bulk_price_editor():
    status_filter   = request.args.get('status', 'processed')
    model_filter    = request.args.get('model', '').strip()
    location_filter = request.args.get('location', '').strip()
    priced_filter   = request.args.get('priced', '').strip()  # 'yes'/'no'/''

    q = ProductInstance.query.join(Product).filter(
        Product.tenant_id == current_user.tenant_id,
        ProductInstance.is_sold == False,
    )
    if status_filter and status_filter != 'all':
        q = q.filter(ProductInstance.status == status_filter)
    if model_filter:
        q = q.filter(Product.model.ilike(f'%{model_filter}%'))
    if location_filter:
        q = q.join(Location, ProductInstance.location_id == Location.id).filter(Location.name == location_filter)
    if priced_filter == 'yes':
        q = q.filter(ProductInstance.asking_price != None)
    elif priced_filter == 'no':
        q = q.filter(ProductInstance.asking_price == None)

    instances = q.order_by(Product.model, Product.cpu, ProductInstance.serial).all()
    locations = Location.query.filter_by(tenant_id=current_user.tenant_id).order_by(Location.name).all()

    return render_template(
        'bulk_price_editor.html',
        instances=instances,
        locations=locations,
        status_filter=status_filter,
        model_filter=model_filter,
        location_filter=location_filter,
        priced_filter=priced_filter,
    )


@stock_bp.route('/bulk_price/save', methods=['POST'])
@login_required
def bulk_price_save():
    data = request.get_json()
    if not data:
        return jsonify({'success': False, 'message': 'No data'}), 400

    instance_ids = data.get('instance_ids', [])
    clear        = data.get('clear', False)
    price_raw    = data.get('price')

    if not instance_ids:
        return jsonify({'success': False, 'message': 'No units selected'}), 400

    if clear:
        price = None
    else:
        try:
            price = float(str(price_raw).strip())
            if price < 0:
                raise ValueError
        except (ValueError, TypeError):
            return jsonify({'success': False, 'message': 'Invalid price value'}), 400

    # Scope to tenant, unsold only
    updated = ProductInstance.query.join(Product).filter(
        ProductInstance.id.in_(instance_ids),
        Product.tenant_id == current_user.tenant_id,
        ProductInstance.is_sold == False,
    ).all()

    if not updated:
        return jsonify({'success': False, 'message': 'No matching units found'}), 404

    for inst in updated:
        inst.asking_price = price
    db.session.commit()

    return jsonify({'success': True, 'count': len(updated), 'price': price})


# ─────────────────────────────────────────────────────────────
# Mobile Scanner — dedicated full-screen scan page
# ─────────────────────────────────────────────────────────────
# Scanner routes (scan_unit, lookup_unit, scan_move_unit, scan_update_status, scan_move)
# moved to routes/scanner.py (scanner_bp)


@stock_bp.route('/api/group_detail')
@login_required
def api_group_detail():
    # Revert to filtering by model and cpu instead of product_id
    model = request.args.get("model", "").strip()
    cpu = request.args.get("cpu", "").strip()
    if not model or not cpu:
        return "Missing model or CPU", 400

    from sqlalchemy import func
    import re
    location_name = request.args.get("location", "").strip()
    ram = request.args.get("ram", "").strip()
    display = request.args.get("display", "").strip()
    stage = request.args.get("stage", "").strip()
    status = request.args.get("status", "").strip()
    gpu = request.args.get("gpu", "").strip()

    # List all products under this tenant (no debug print)
    all_products = Product.query.filter(Product.tenant_id == current_user.tenant_id).all()

    # Query all instances for this tenant, unsold
    instances = ProductInstance.query.join(Product).outerjoin(Location, Location.id == ProductInstance.location_id).filter(
        Product.tenant_id == current_user.tenant_id,
        ProductInstance.is_sold == False
    )
    # Exclude reserved units (like under_process)
    reserved_order = aliased(CustomerOrderTracking)
    instances = instances.outerjoin(
        reserved_order,
        reserved_order.product_instance_id == ProductInstance.id
    ).filter(or_(reserved_order.id == None, reserved_order.status != 'reserved'))
    # Optionally reapply other filters here (ram, display, etc.)
    if ram:
        instances = instances.filter(Product.ram == ram)
    if display:
        instances = instances.filter(Product.display == display)
    if gpu:
        instances = instances.filter(Product.gpu1 == gpu)
    if stage:
        instances = instances.filter(ProductInstance.process_stage == stage)
    # Only apply status filter if 'status' is explicitly present in query params
    if status and status.strip().lower() not in ("", "all", "none"):
        instances = instances.filter(ProductInstance.status == status)
    if location_name:
        instances = instances.filter(func.lower(Location.name) == location_name.lower())

    # --- Robust model/cpu filtering using SQL, whitespace/symbol normalization ---
    def _clean(s):
        s = (s or "")
        s = s.replace("®", "").replace("™", "")
        s = s.strip().lower()
        # normalize any whitespace (tabs/newlines/multiple spaces) to single spaces
        s = re.sub(r"\s+", " ", s)
        return s
    model_clean = _clean(model)
    cpu_clean = _clean(cpu)
    model_sql = func.lower(func.regexp_replace(func.trim(Product.model), r'\s+', ' ', 'g'))
    cpu_sql = func.lower(func.regexp_replace(func.trim(Product.cpu), r'\s+', ' ', 'g'))
    instances = instances.filter(model_sql == model_clean, cpu_sql == cpu_clean)

    instances = instances.options(
        joinedload(ProductInstance.product),
        joinedload(ProductInstance.location)
    ).all()

    from inventory_flask_app.models import TenantSettings
    tenant_settings = TenantSettings.query.filter_by(tenant_id=current_user.tenant_id).all()
    settings_dict = {s.key: s.value for s in tenant_settings}
    column_order = settings_dict.get("column_order_instance_table")

    default_columns = [
        "asset", "serial", "Item Name", "make", "model", "display", "cpu", "ram",
        "gpu1", "gpu2", "Grade", "location", "status", "process_stage",
        "team_assigned", "shelf_bin", "is_sold", "label", "action"
    ]

    _alias_map = {
        # Old API names
        "model_number": "model",
        "product": "Item Name",
        "processor": "cpu",
        "video_card": "gpu1",
        "resolution": "display",
        "screen_size": "display",
        # Case/format variants saved by admin column reorder
        "item_name": "Item Name",
        "item name": "Item Name",
        "grade": "Grade",
        "team": "team_assigned",
        "gpu": "gpu1",
    }
    if column_order:
        seen = set()
        existing = []
        for col in column_order.split(","):
            col = _alias_map.get(col.strip(), col.strip())
            if col and col not in seen:
                seen.add(col)
                existing.append(col)
        for col in default_columns:
            if col not in seen:
                seen.add(col)
                existing.append(col)
        column_order = ",".join(existing)
    else:
        column_order = ",".join(default_columns)

    settings_dict["column_order_instance_table"] = column_order

    from inventory_flask_app.models import CustomField, CustomFieldValue
    list_custom_fields = CustomField.query.filter_by(
        tenant_id=current_user.tenant_id, show_in_list=True
    ).order_by(CustomField.sort_order).all()
    instance_ids = [i.id for i in instances]
    cf_values_map = {}
    if instance_ids and list_custom_fields:
        for cfv in CustomFieldValue.query.filter(
            CustomFieldValue.instance_id.in_(instance_ids)
        ).all():
            cf_values_map.setdefault(cfv.instance_id, {})[cfv.field_id] = cfv.value

    # Sort unit list
    unit_sort     = request.args.get('sort', 'serial')
    unit_sort_dir = request.args.get('sort_dir', 'asc')
    unit_reverse  = (unit_sort_dir == 'desc')
    _now_utc = datetime.utcnow()
    _sort_key = {
        'serial':        lambda i: (i.serial or '').lower(),
        'status':        lambda i: (i.status or '').lower(),
        'process_stage': lambda i: (i.process_stage or '').lower(),
        'location':      lambda i: (i.location.name if i.location else '').lower(),
        'grade':         lambda i: (i.product.grade if i.product else '').lower(),
        'age_days':      lambda i: ((_now_utc - i.created_at).days if i.created_at else 0),
        'asking_price':  lambda i: (i.asking_price or 0),
    }.get(unit_sort, lambda i: (i.serial or '').lower())
    instances = sorted(instances, key=_sort_key, reverse=unit_reverse)

    return render_template(
        "group_view.html",
        instances=instances,
        settings=settings_dict,
        location_name=location_name,
        model=model,
        cpu=cpu,
        list_custom_fields=list_custom_fields,
        cf_values_map=cf_values_map,
        unit_sort=unit_sort,
        unit_sort_dir=unit_sort_dir,
    )

# Stock Intake hub
@stock_bp.route('/stock_intake')
@login_required
def stock_intake():
    from inventory_flask_app.models import PurchaseOrderItem
    pos = PurchaseOrder.query.filter_by(
        tenant_id=current_user.tenant_id
    ).options(
        joinedload(PurchaseOrder.vendor),
        joinedload(PurchaseOrder.location),
    ).order_by(PurchaseOrder.created_at.desc()).limit(60).all()

    po_ids = [po.id for po in pos]
    stats_map = {}
    if po_ids:
        for row in db.session.query(
            PurchaseOrderItem.po_id,
            func.count(PurchaseOrderItem.id).label('total'),
            func.sum(db.case((PurchaseOrderItem.status == 'received', 1), else_=0)).label('received'),
        ).filter(PurchaseOrderItem.po_id.in_(po_ids)).group_by(PurchaseOrderItem.po_id).all():
            stats_map[row.po_id] = {'total': row.total, 'received': row.received or 0}

    return render_template('stock_intake.html', pos=pos, stats_map=stats_map)

@stock_bp.route('/po_template_download')
@login_required
def po_template_download():
    from openpyxl import Workbook
    from flask import send_file
    wb = Workbook()
    ws = wb.active
    ws.title = 'PO Import'
    ws.append(['serial', 'asset', 'item_name', 'make', 'model', 'cpu', 'ram', 'grade', 'display', 'gpu1', 'gpu2', 'disk1size'])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, download_name='po_import_template.xlsx', as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@stock_bp.route('/purchase_order/create', methods=['GET', 'POST'])
@login_required
def create_purchase_order():
    if request.method == 'POST':
        po_number = (request.form.get('po_number') or '').strip()
        vendor_id  = request.form.get('vendor_id')
        location_id = request.form.get('location_id')
        notes_val  = (request.form.get('notes') or '').strip()
        file       = request.files.get('file')

        if not po_number or not vendor_id or not file or file.filename == '':
            flash('PO Number, Vendor and Excel file are required.', 'danger')
            return redirect(url_for('stock_bp.create_purchase_order'))

        if PurchaseOrder.query.filter_by(po_number=po_number, tenant_id=current_user.tenant_id).first():
            flash(f"PO Number '{po_number}' already exists.", 'danger')
            return redirect(url_for('stock_bp.create_purchase_order'))

        import pandas as pd
        try:
            df = pd.read_excel(file)
        except Exception as e:
            flash(f'Could not read Excel file: {e}', 'danger')
            return redirect(url_for('stock_bp.create_purchase_order'))

        # Normalise column names: strip, lowercase, spaces→underscores
        df.columns = [c.strip().lower().replace(' ', '_') for c in df.columns]

        # Common alias mappings from vendor spreadsheets
        df.rename(columns={
            'serial_number':  'serial',
            'sn':             'serial',
            'asset_tag':      'asset',
            'asset_number':   'asset',
            'product_name':   'item_name',
            'item':           'item_name',
            'description':    'item_name',
            'manufacturer':   'make',
            'brand':          'make',
            'screen_size':    'display',
            'hdd':            'disk1size',
            'disk':           'disk1size',
            'storage':        'disk1size',
            'processor':      'cpu',
            'memory':         'ram',
        }, inplace=True)

        # Serial is the only truly required column
        if 'serial' not in df.columns:
            flash("Excel must include a 'serial' column (or 'serial_number' / 'sn').", 'danger')
            return redirect(url_for('stock_bp.create_purchase_order'))

        # Drop fully empty rows
        df.dropna(how='all', inplace=True)

        def _clean(v):
            s = str(v).strip() if v is not None else ''
            return '' if s.lower() in ('nan', 'none', 'n/a', '-') else s

        items_added = 0
        skipped = 0
        po = PurchaseOrder(
            po_number=po_number,
            vendor_id=int(vendor_id) if vendor_id else None,
            location_id=int(location_id) if location_id else None,
            tenant_id=current_user.tenant_id,
            status='pending',
            notes=notes_val or None,
        )
        db.session.add(po)
        db.session.flush()  # get po.id

        from inventory_flask_app.models import PurchaseOrderItem
        for _, row in df.iterrows():
            serial = _clean(row.get('serial'))
            if not serial:
                skipped += 1
                continue

            poi = PurchaseOrderItem(
                po_id=po.id,
                tenant_id=current_user.tenant_id,
                serial=serial,
                asset_tag=_clean(row.get('asset')) or None,
                item_name=_clean(row.get('item_name')) or None,
                make=_clean(row.get('make')) or None,
                model=_clean(row.get('model')) or None,
                display=_clean(row.get('display')) or None,
                cpu=_clean(row.get('cpu')) or None,
                ram=_clean(row.get('ram')) or None,
                gpu1=_clean(row.get('gpu1')) or None,
                gpu2=_clean(row.get('gpu2')) or None,
                grade=_clean(row.get('grade')) or None,
                disk1size=_clean(row.get('disk1size')) or None,
                location_id=int(location_id) if location_id else None,
                status='expected',
            )
            db.session.add(poi)
            items_added += 1

        if items_added == 0:
            db.session.rollback()
            flash('No valid rows found in Excel (serial column was empty for all rows).', 'warning')
            return redirect(url_for('stock_bp.create_purchase_order'))

        # Also keep expected_serials for backward compat with old session-based code paths
        serials_list = [_clean(r.get('serial')) for _, r in df.iterrows() if _clean(r.get('serial'))]
        po.expected_serials = ','.join(serials_list)
        db.session.commit()

        msg = f'PO {po.po_number} created with {items_added} item(s).'
        if skipped:
            msg += f' {skipped} row(s) skipped (no serial).'
        flash(msg, 'success')
        return redirect(url_for('stock_bp.stock_receiving_select'))

    vendors   = Vendor.query.filter_by(tenant_id=current_user.tenant_id).all()
    locations = Location.query.filter_by(tenant_id=current_user.tenant_id).all()
    return render_template('create_purchase_order.html', vendors=vendors, locations=locations)


@stock_bp.route('/purchase_order/<int:po_id>')
@login_required
def view_purchase_order(po_id):
    """PO detail — shows all expected items and their receive status."""
    from inventory_flask_app.models import PurchaseOrderItem
    po = PurchaseOrder.query.filter_by(
        id=po_id, tenant_id=current_user.tenant_id
    ).first_or_404()
    items = PurchaseOrderItem.query.filter_by(po_id=po.id).order_by(
        PurchaseOrderItem.status, PurchaseOrderItem.serial
    ).all()
    total    = len(items)
    received = sum(1 for i in items if i.status == 'received')
    missing  = sum(1 for i in items if i.status == 'missing')
    extra    = sum(1 for i in items if i.status == 'extra')
    return render_template(
        'view_purchase_order.html',
        po=po, items=items,
        total=total, received=received, missing=missing, extra=extra,
    )


# ─────────────────────────────────────────────────────────────
# Stock Receiving — rebuilt to use DB-backed PurchaseOrderItems
# ─────────────────────────────────────────────────────────────

@stock_bp.route('/stock_receiving/select', methods=['GET', 'POST'])
@login_required
def stock_receiving_select():
    # Direct-link shortcut: ?po_id=X skips the select form
    if request.method == 'GET' and request.args.get('po_id'):
        po = PurchaseOrder.query.filter_by(
            id=request.args.get('po_id', type=int),
            tenant_id=current_user.tenant_id,
        ).first()
        if po and po.status in ('pending', 'partial'):
            session['po_id'] = po.id
            session['scanned'] = []
            session.modified = True
            return redirect(url_for('stock_bp.stock_receiving_scan'))

    # Show pending + partial POs
    po_list = PurchaseOrder.query.filter(
        PurchaseOrder.status.in_(['pending', 'partial']),
        PurchaseOrder.tenant_id == current_user.tenant_id
    ).options(
        joinedload(PurchaseOrder.vendor),
        joinedload(PurchaseOrder.location),
    ).order_by(PurchaseOrder.created_at.desc()).all()

    # Build per-PO item stats
    from inventory_flask_app.models import PurchaseOrderItem
    po_ids = [po.id for po in po_list]
    stats_map = {}
    if po_ids:
        for row in db.session.query(
            PurchaseOrderItem.po_id,
            func.count(PurchaseOrderItem.id).label('total'),
            func.sum(db.case((PurchaseOrderItem.status == 'received', 1), else_=0)).label('received'),
        ).filter(PurchaseOrderItem.po_id.in_(po_ids)).group_by(PurchaseOrderItem.po_id).all():
            stats_map[row.po_id] = {'total': row.total, 'received': row.received or 0}

    if request.method == 'POST':
        po_id = request.form.get('po_id')
        po = PurchaseOrder.query.filter_by(id=po_id, tenant_id=current_user.tenant_id).first()
        if not po:
            flash('Purchase Order not found.', 'danger')
            return redirect(url_for('stock_bp.stock_receiving_select'))
        # Store only the PO id in session; all spec data comes from DB
        session['po_id'] = po.id
        session['scanned'] = []
        session.modified = True
        return redirect(url_for('stock_bp.stock_receiving_scan'))

    return render_template('stock_receiving_select.html', po_list=po_list, stats_map=stats_map)


@stock_bp.route('/stock_receiving/scan', methods=['GET'])
@login_required
def stock_receiving_scan():
    po_id = session.get('po_id')
    if not po_id:
        flash('Select a PO first.', 'warning')
        return redirect(url_for('stock_bp.stock_receiving_select'))

    po = PurchaseOrder.query.filter_by(id=po_id, tenant_id=current_user.tenant_id).first_or_404()
    from inventory_flask_app.models import PurchaseOrderItem
    items = PurchaseOrderItem.query.filter_by(po_id=po.id).all()

    # Build lookup maps from DB items
    serial_map   = {i.serial.strip().lower(): i for i in items}
    asset_map    = {i.asset_tag.strip().lower(): i for i in items if i.asset_tag}

    scanned = session.get('scanned', [])
    scanned_rows = _build_scan_rows(scanned, serial_map, asset_map)

    matched_count = sum(1 for r in scanned_rows if r['match'] == 'matched')
    extra_count   = sum(1 for r in scanned_rows if r['match'] == 'extra')
    total_expected = len(items)
    missing_count = total_expected - matched_count

    matched_serials = {r['serial'].strip().lower() for r in scanned_rows if r['match'] == 'matched'}

    return render_template(
        'stock_receiving_scan.html',
        po=po,
        items=items,
        scanned_rows=scanned_rows,
        total_expected=total_expected,
        matched_count=matched_count,
        extra_count=extra_count,
        missing_count=missing_count,
        matched_serials=matched_serials,
    )


@csrf.exempt
@stock_bp.route('/stock_receiving/scan_item', methods=['POST'])
@login_required
def stock_receiving_scan_item():
    """AJAX endpoint: add one scan to session, return updated counts + row data."""
    po_id = session.get('po_id')
    if not po_id:
        return jsonify({'error': 'No PO in session'}), 400

    po = PurchaseOrder.query.filter_by(id=po_id, tenant_id=current_user.tenant_id).first()
    if not po:
        return jsonify({'error': 'PO not found'}), 404

    data = request.get_json(silent=True) or {}
    value = (data.get('value') or '').strip()
    if not value:
        return jsonify({'error': 'Empty scan value'}), 400

    scanned = session.get('scanned', [])
    if value in scanned:
        return jsonify({'duplicate': True, 'value': value})

    from inventory_flask_app.models import PurchaseOrderItem
    items = PurchaseOrderItem.query.filter_by(po_id=po.id).all()
    serial_map = {i.serial.strip().lower(): i for i in items}
    asset_map  = {i.asset_tag.strip().lower(): i for i in items if i.asset_tag}

    v_lower = value.strip().lower()
    if v_lower in serial_map:
        poi  = serial_map[v_lower]
        match = 'matched'
    elif v_lower in asset_map:
        poi  = asset_map[v_lower]
        match = 'matched'
    else:
        poi  = None
        match = 'extra'

    scanned.append(value)
    session['scanned'] = scanned
    session.modified = True

    scanned_rows = _build_scan_rows(scanned, serial_map, asset_map)
    matched_count = sum(1 for r in scanned_rows if r['match'] == 'matched')
    extra_count   = sum(1 for r in scanned_rows if r['match'] == 'extra')
    missing_count = len(items) - matched_count

    row_data = {
        'value':     value,
        'match':     match,
        'serial':    poi.serial    if poi else value,
        'asset_tag': poi.asset_tag if poi else '',
        'model':     poi.model     if poi else '',
        'make':      poi.make      if poi else '',
        'cpu':       poi.cpu       if poi else '',
        'ram':       poi.ram       if poi else '',
    }
    return jsonify({
        'ok':            True,
        'row':           row_data,
        'matched_count': matched_count,
        'extra_count':   extra_count,
        'missing_count': missing_count,
        'total_scanned': len(scanned),
    })


@csrf.exempt
@stock_bp.route('/stock_receiving/scan_reset', methods=['POST'])
@login_required
def stock_receiving_scan_reset():
    """AJAX: clear session scanned list."""
    session['scanned'] = []
    session.modified = True
    return jsonify({'ok': True})


def _build_scan_rows(scanned, serial_map, asset_map):
    """Build display rows from a list of scanned values against DB item maps."""
    rows = []
    for v in scanned:
        v_lower = v.strip().lower()
        poi = serial_map.get(v_lower) or asset_map.get(v_lower)
        rows.append({
            'value':     v,
            'match':     'matched' if poi else 'extra',
            'serial':    poi.serial    if poi else v,
            'asset_tag': poi.asset_tag if poi else '',
            'model':     poi.model     if poi else '',
            'make':      poi.make      if poi else '',
            'cpu':       poi.cpu       if poi else '',
            'ram':       poi.ram       if poi else '',
        })
    return rows


@stock_bp.route('/stock_receiving/summary')
@login_required
def stock_receiving_summary():
    po_id = session.get('po_id')
    if not po_id:
        flash('Select a PO first.', 'warning')
        return redirect(url_for('stock_bp.stock_receiving_select'))

    po = PurchaseOrder.query.filter_by(id=po_id, tenant_id=current_user.tenant_id).first_or_404()
    from inventory_flask_app.models import PurchaseOrderItem
    items = PurchaseOrderItem.query.filter_by(po_id=po.id).all()

    serial_map = {i.serial.strip().lower(): i for i in items}
    asset_map  = {i.asset_tag.strip().lower(): i for i in items if i.asset_tag}
    scanned    = session.get('scanned', [])

    # Classify each scanned value
    scanned_matched_keys = set()
    extra_values = []
    for v in scanned:
        v_lower = v.strip().lower()
        if v_lower in serial_map:
            scanned_matched_keys.add(v_lower)
        elif v_lower in asset_map:
            scanned_matched_keys.add(asset_map[v_lower].serial.strip().lower())
        else:
            extra_values.append(v)

    # Annotate each PO item
    table_rows = []
    for item in items:
        key = item.serial.strip().lower()
        if key in scanned_matched_keys:
            match_status = 'matched'
        elif item.status == 'received':
            match_status = 'previously_received'
        else:
            match_status = 'missing'
        existing_id = get_instance_id(item.serial)
        table_rows.append({
            'poi':          item,
            'match_status': match_status,
            'instance_id':  existing_id,
        })

    # Extra rows (scanned but not on PO)
    for v in extra_values:
        table_rows.append({
            'poi':          None,
            'match_status': 'extra',
            'instance_id':  get_instance_id(v),
            'extra_value':  v,
        })

    matched_count  = sum(1 for r in table_rows if r['match_status'] == 'matched')
    missing_count  = sum(1 for r in table_rows if r['match_status'] == 'missing')
    extra_count    = len(extra_values)
    prev_count     = sum(1 for r in table_rows if r['match_status'] == 'previously_received')
    locations      = Location.query.filter_by(tenant_id=current_user.tenant_id).all()

    return render_template(
        'stock_receiving_summary.html',
        po=po,
        table_rows=table_rows,
        matched_count=matched_count,
        missing_count=missing_count,
        extra_count=extra_count,
        prev_count=prev_count,
        locations=locations,
    )


def export_csv(data, columns, filename):
    output = StringIO()
    writer = csv.DictWriter(output, fieldnames=columns)
    writer.writeheader()
    for row in data:
        writer.writerow({col: row.get(col, '') for col in columns})
    response = make_response(output.getvalue())
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    response.headers['Content-Type'] = 'text/csv'
    return response


@stock_bp.route('/stock_receiving/export/<category>')
@login_required
def stock_receiving_export(category):
    po_id = session.get('po_id')
    if not po_id:
        flash('No active PO session.', 'warning')
        return redirect(url_for('stock_bp.stock_receiving_select'))

    from inventory_flask_app.models import PurchaseOrderItem
    items = PurchaseOrderItem.query.filter_by(po_id=po_id).all()
    scanned = set(v.strip().lower() for v in session.get('scanned', []))
    serial_map = {i.serial.strip().lower(): i for i in items}
    asset_map  = {i.asset_tag.strip().lower(): i for i in items if i.asset_tag}
    matched_keys = set()
    for v in scanned:
        if v in serial_map:
            matched_keys.add(v)
        elif v in asset_map:
            matched_keys.add(asset_map[v].serial.strip().lower())

    export_rows = []
    for item in items:
        key = item.serial.strip().lower()
        if category == 'matched' and key in matched_keys:
            export_rows.append({'serial': item.serial, 'asset': item.asset_tag or '', 'model': item.model or '', 'cpu': item.cpu or '', 'ram': item.ram or '', 'status': 'matched'})
        elif category == 'missing' and key not in matched_keys and item.status != 'received':
            export_rows.append({'serial': item.serial, 'asset': item.asset_tag or '', 'model': item.model or '', 'cpu': item.cpu or '', 'ram': item.ram or '', 'status': 'missing'})

    if not export_rows:
        flash(f"No rows found for category '{category}'.", 'info')
        return redirect(url_for('stock_bp.stock_receiving_summary'))

    return export_csv(export_rows, ['serial', 'asset', 'model', 'cpu', 'ram', 'status'], f'{category}_receiving_{po_id}.csv')


@stock_bp.route('/stock_receiving/confirm', methods=['POST'])
@login_required
def stock_receiving_confirm():
    po_id       = session.get('po_id')
    scanned     = session.get('scanned', [])
    status_choice  = request.form.get('status_choice')
    location_id    = request.form.get('location_choice')

    if not po_id:
        flash('Session expired. Select a PO again.', 'danger')
        return redirect(url_for('stock_bp.stock_receiving_select'))
    if not scanned:
        flash('Nothing was scanned.', 'warning')
        return redirect(url_for('stock_bp.stock_receiving_scan'))
    if not status_choice:
        flash('Please select an inventory status before importing.', 'danger')
        return redirect(url_for('stock_bp.stock_receiving_summary'))

    allowed_statuses = {'unprocessed', 'under_process', 'processed'}
    if status_choice not in allowed_statuses:
        flash('Invalid status selected.', 'danger')
        return redirect(url_for('stock_bp.stock_receiving_summary'))

    po = PurchaseOrder.query.filter_by(id=po_id, tenant_id=current_user.tenant_id).first_or_404()

    from inventory_flask_app.models import PurchaseOrderItem, POImportLog
    from inventory_flask_app.utils.utils import upsert_instance

    items      = PurchaseOrderItem.query.filter_by(po_id=po.id).all()
    serial_map = {i.serial.strip().lower(): i for i in items}
    asset_map  = {i.asset_tag.strip().lower(): i for i in items if i.asset_tag}
    now_ts     = get_now_for_tenant()

    result_rows   = []
    created_count = 0
    updated_count = 0
    skipped_count = 0
    failed_count  = 0

    for v in scanned:
        v_lower = v.strip().lower()
        poi = serial_map.get(v_lower) or asset_map.get(v_lower)
        if not poi:
            # extra — not on PO; skip silently
            continue

        serial = poi.serial.strip()
        spec_data = {
            'item_name': poi.item_name or poi.model or '',
            'make':      poi.make or '',
            'model':     poi.model or '',
            'cpu':       poi.cpu or '',
            'ram':       poi.ram or '',
            'display':   getattr(poi, 'display', '') or '',
            'gpu1':      getattr(poi, 'gpu1', '') or '',
            'gpu2':      getattr(poi, 'gpu2', '') or '',
            'grade':     poi.grade or '',
            'disk1size': getattr(poi, 'disk1size', '') or '',
            'asset':     poi.asset_tag or '',
        }

        def _create_fn(poi=poi, po=po, loc=location_id):
            return _find_or_create_product(poi, po.vendor_id, current_user.tenant_id, loc)

        try:
            sp = db.session.begin_nested()
            outcome, instance, changes = upsert_instance(
                serial=serial,
                spec_data=spec_data,
                tenant_id=current_user.tenant_id,
                location_id=int(location_id) if location_id else None,
                vendor_id=po.vendor_id,
                po_id=po.id,
                status=status_choice,
                moved_by_id=current_user.id,
                create_product_fn=_create_fn,
            )
            sp.commit()
            result_rows.append({
                'serial': serial, 'outcome': outcome,
                'instance': instance, 'changes': changes,
            })
            # Mark POI as received for created and updated units
            if outcome in ('created', 'updated'):
                poi.status      = 'received'
                poi.received_at = now_ts
                if outcome == 'created':
                    created_count += 1
                else:
                    updated_count += 1
            else:
                skipped_count += 1

        except Exception as exc:
            sp.rollback()
            logger.exception("stock_receiving_confirm: serial=%s failed: %s", serial, exc)
            failed_count += 1
            result_rows.append({
                'serial': serial, 'outcome': 'failed',
                'instance': None, 'changes': {}, 'error': str(exc),
            })

    # Update PO status: partial if some items still expected, received if all done
    remaining = PurchaseOrderItem.query.filter_by(po_id=po.id, status='expected').count()
    po.status = 'received' if remaining == 0 else 'partial'

    received_count = created_count + updated_count
    log = POImportLog(po_id=po.id, user_id=current_user.id, status=status_choice, quantity=received_count)
    db.session.add(log)
    db.session.commit()

    session.pop('scanned', None)
    session.pop('po_id', None)
    session.modified = True

    locations = Location.query.filter_by(tenant_id=current_user.tenant_id).all()
    return render_template(
        'stock_receiving_result.html',
        po=po,
        result_rows=result_rows,
        created_count=created_count,
        updated_count=updated_count,
        skipped_count=skipped_count,
        failed_count=failed_count,
        status_choice=status_choice,
        locations=locations,
    )


def _find_or_create_product(poi, vendor_id, tenant_id, location_id):
    """Find a matching product by key spec fields, or create a new one.
    Uses (model, cpu, ram, make, grade, tenant_id) as the match key — tolerates
    null optional fields so we don't create a new product for every received unit.
    """
    filters = [Product.tenant_id == tenant_id]
    if poi.model:
        filters.append(Product.model == poi.model)
    if poi.cpu:
        filters.append(Product.cpu == poi.cpu)
    if poi.ram:
        filters.append(Product.ram == poi.ram)
    if poi.make:
        filters.append(Product.make == poi.make)

    product = Product.query.filter(*filters).first() if len(filters) > 1 else None

    if not product:
        product = Product(
            item_name=poi.item_name or poi.model or 'Received Unit',
            make=poi.make,
            model=poi.model,
            display=poi.display,
            cpu=poi.cpu,
            ram=poi.ram,
            gpu1=poi.gpu1,
            gpu2=poi.gpu2,
            grade=poi.grade,
            disk1size=poi.disk1size,
            vendor_id=vendor_id,
            tenant_id=tenant_id,
            location_id=int(location_id) if location_id else None,
            created_at=get_now_for_tenant(),
        )
        db.session.add(product)
        db.session.flush()

    return product


@stock_bp.route('/api/model_suggestions')
@login_required
def model_suggestions():
    q = request.args.get('q', '').strip()
    if not q or len(q) < 2:
        return jsonify([])
    # Query distinct models that contain the substring, case-insensitive
    models = (
        db.session.query(Product.model)
        .filter(
            Product.tenant_id == current_user.tenant_id,
            Product.model.ilike(f"%{q}%")
        )
        .distinct()
        .order_by(Product.model)
        .limit(15)
        .all()
    )
    # Return just the list of model names
    return jsonify([m[0] for m in models if m[0]])

# Batch label printing route
@stock_bp.route('/print_labels_batch', methods=['POST'])
@login_required
def print_labels_batch():
    from sqlalchemy import or_, and_
    ids = request.form.getlist('instance_ids')
    if not ids:
        flash("No items selected for label printing.", "warning")
        return redirect(request.referrer or url_for('stock_bp.under_process'))

    # Support both integer IDs (from group_view unit-level checkboxes)
    # and model|||cpu keys (from instance_table grouped-level checkboxes)
    grouped_keys = [v for v in ids if "|||" in v]
    exact_ids    = [int(v) for v in ids if v.isdigit()]

    base_q = ProductInstance.query.join(Product).filter(
        Product.tenant_id == current_user.tenant_id
    )
    if grouped_keys and not exact_ids:
        filters = []
        for key in grouped_keys:
            m, c = key.split("|||", 1)
            filters.append(and_(Product.model == m.strip(), Product.cpu == c.strip()))
        base_q = base_q.filter(or_(*filters))
    elif exact_ids:
        base_q = base_q.filter(ProductInstance.id.in_(exact_ids))
    else:
        flash("No valid items selected for label printing.", "warning")
        return redirect(request.referrer or url_for('stock_bp.under_process'))

    instances = base_q.all()
    # Generate QR codes for each
    batch_labels = []
    for instance in instances:
        qr_data = {
            "asset": instance.asset if instance else "",
            "serial": instance.serial if instance else "",
            "item_name": instance.product.item_name if instance.product else "",
            "make": instance.product.make if instance.product else "",
            "model": instance.product.model if instance.product else "",
            "display": instance.product.display if instance.product else "",
            "cpu": instance.product.cpu if instance.product else "",
            "ram": instance.product.ram if instance.product else "",
            "gpu1": instance.product.gpu1 if instance.product else "",
            "gpu2": instance.product.gpu2 if instance.product else "",
            "grade": instance.product.grade if instance.product else "",
            "disk1size": instance.product.disk1size if instance.product else "",
        }
        qr = qrcode.QRCode(version=1, box_size=10, border=2)
        qr.add_data(qr_data)
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")
        buffered = BytesIO()
        img.save(buffered, format="PNG")
        qr_b64 = base64.b64encode(buffered.getvalue()).decode()
        batch_labels.append({
            "instance": instance,
            "qr_b64": qr_b64,
            "printed_time": get_now_for_tenant(),
        })
    return render_template("batch_print_labels.html", batch_labels=batch_labels)


# ── Bulk status change ────────────────────────────────────────────────────────
@stock_bp.route('/bulk_status_change', methods=['POST'])
@login_required
def bulk_status_change():
    from sqlalchemy import or_, and_
    from inventory_flask_app.models import ProductProcessLog

    new_status = request.form.get('new_status', '').strip()
    valid_statuses = {'unprocessed', 'under_process', 'processed', 'idle', 'disputed'}
    if new_status not in valid_statuses:
        flash('Invalid status selected.', 'danger')
        return redirect(request.referrer or url_for('stock_bp.under_process'))

    ids = request.form.getlist('instance_ids')
    if not ids:
        flash('No units selected.', 'warning')
        return redirect(request.referrer or url_for('stock_bp.under_process'))

    grouped_keys = [v for v in ids if '|||' in v]
    exact_ids    = [int(v) for v in ids if v.isdigit()]

    base_q = ProductInstance.query.join(Product).filter(
        Product.tenant_id == current_user.tenant_id
    )
    if grouped_keys and not exact_ids:
        filters = []
        for key in grouped_keys:
            m, c = key.split('|||', 1)
            filters.append(and_(Product.model == m.strip(), Product.cpu == c.strip()))
        base_q = base_q.filter(or_(*filters))
    elif exact_ids:
        base_q = base_q.filter(ProductInstance.id.in_(exact_ids))
    else:
        flash('No valid units selected.', 'warning')
        return redirect(request.referrer or url_for('stock_bp.under_process'))

    instances = base_q.all()
    updated = 0
    now = get_now_for_tenant()
    for inst in instances:
        old_status = inst.status
        inst.status = new_status
        inst.updated_at = now
        log = ProductProcessLog(
            product_instance_id=inst.id,
            from_stage=old_status,
            to_stage=new_status,
            moved_by=current_user.id,
            moved_at=now,
            action='bulk_status_change',
            note=f'Bulk changed from {old_status} to {new_status}',
        )
        db.session.add(log)
        updated += 1

    db.session.commit()
    flash(f'Updated {updated} unit(s) to "{new_status}".', 'success')
    return redirect(request.referrer or url_for('stock_bp.under_process'))


# ── Bulk move to bin ──────────────────────────────────────────────────────────
@stock_bp.route('/bulk_move_to_bin', methods=['POST'])
@login_required
def bulk_move_to_bin():
    from sqlalchemy import or_, and_

    bin_id = request.form.get('bin_id', type=int)
    ids    = request.form.getlist('instance_ids')
    if not ids:
        flash('No units selected.', 'warning')
        return redirect(request.referrer or url_for('stock_bp.under_process'))

    target_bin = None
    if bin_id:
        target_bin = Bin.query.filter_by(id=bin_id, tenant_id=current_user.tenant_id).first()

    grouped_keys = [v for v in ids if '|||' in v]
    exact_ids    = [int(v) for v in ids if v.isdigit()]

    base_q = ProductInstance.query.join(Product).filter(
        Product.tenant_id == current_user.tenant_id
    )
    if grouped_keys and not exact_ids:
        filters = []
        for key in grouped_keys:
            m, c = key.split('|||', 1)
            filters.append(and_(Product.model == m.strip(), Product.cpu == c.strip()))
        base_q = base_q.filter(or_(*filters))
    elif exact_ids:
        base_q = base_q.filter(ProductInstance.id.in_(exact_ids))
    else:
        flash('No valid units selected.', 'warning')
        return redirect(request.referrer or url_for('stock_bp.under_process'))

    instances = base_q.all()
    updated = 0
    for inst in instances:
        if target_bin:
            inst.bin_id      = target_bin.id
            inst.shelf_bin   = target_bin.name
            inst.location_id = target_bin.location_id
        else:
            inst.bin_id    = None
            inst.shelf_bin = None
        updated += 1

    db.session.commit()
    dest = target_bin.name if target_bin else 'no bin'
    flash(f'Moved {updated} unit(s) to {dest}.', 'success')
    return redirect(request.referrer or url_for('stock_bp.under_process'))


@stock_bp.route('/under_process', methods=['GET', 'POST'])
@login_required
def under_process():
    status_filter = request.args.get('status')
    model_filter = request.args.get('model')
    processor_filter = request.args.get('processor')
    serial_search = request.args.get('serial_search', '').strip()
    stage_filter = request.args.get('stage')
    team_filter = request.args.get('team')
    # location_id = request.args.get('location_id')
    location_filter = request.args.get('location', '').strip()
    bin_search = request.args.get('bin_search', '').strip()
    # New RAM and Disk filters
    ram_filter = request.args.get('ram')
    disk_filter = request.args.get('disk1size')

    # Improved status filter logic: always use .filter(), never .filter_by() for status
    if not status_filter or status_filter == 'all':
        query = ProductInstance.query
    else:
        query = ProductInstance.query
        if status_filter == 'unprocessed':
            query = query.filter(ProductInstance.status == 'unprocessed', ProductInstance.is_sold == False)
        elif status_filter == 'under_process':
            query = query.filter(ProductInstance.status == 'under_process', ProductInstance.is_sold == False)
        elif status_filter == 'processed':
            query = query.filter(ProductInstance.status == 'processed', ProductInstance.is_sold == False)
        elif status_filter == 'disputed':
            query = query.filter(ProductInstance.status == 'disputed', ProductInstance.is_sold == False)
        else:
            query = query.filter(ProductInstance.status == status_filter)

    # Eager load related product and location data for performance
    query = query.options(
        joinedload(ProductInstance.product),
        joinedload(ProductInstance.location)
    )

    # Low stock filter (for /stock/under_process?low_stock=1)
    low_stock = request.args.get('low_stock')
    if low_stock:
        low_stock_product_ids = (
            db.session.query(ProductInstance.product_id)
            .join(Product)
            .filter(
                Product.tenant_id == current_user.tenant_id,
                ProductInstance.is_sold == False
            )
            .group_by(ProductInstance.product_id)
            .having(func.count(ProductInstance.id) <= 3)
            .subquery()
        )
        query = query.filter(ProductInstance.product_id.in_(db.session.query(low_stock_product_ids.c.product_id)))

    # Join Product only once if any filter is present (or always, to simplify filter logic)
    if model_filter or processor_filter or ram_filter or disk_filter or True:
        query = query.join(Product)
    if model_filter:
        query = query.filter(Product.model.ilike(f"%{model_filter}%"))
    if processor_filter:
        query = query.filter(Product.cpu == processor_filter)
    # RAM and Disk filter logic (re-added after model/processor filter)
    if ram_filter:
        query = query.filter(Product.ram == ram_filter)
    if disk_filter:
        query = query.filter(Product.disk1size == disk_filter)
    video_card_filter = request.args.get('video_card', '').strip()
    if video_card_filter:
        query = query.filter(
            or_(Product.gpu1.ilike(f'%{video_card_filter}%'), Product.gpu2.ilike(f'%{video_card_filter}%'))
        )
    if serial_search:
        query = query.filter(or_(
            ProductInstance.serial.ilike(f"%{serial_search}%"),
            ProductInstance.asset.ilike(f"%{serial_search}%")
        ))
    if stage_filter:
        query = query.filter(ProductInstance.process_stage == stage_filter)
    if team_filter:
        query = query.filter(ProductInstance.team_assigned.ilike(f"%{team_filter}%"))
    # Location name filter (new)
    if location_filter:
        query = query.join(ProductInstance.location).filter(Location.name.ilike(location_filter))
    # If you want to keep supporting location_id filter (by id), you may add:
    # if location_id:
    #     query = query.filter(ProductInstance.location_id == int(location_id))
    if bin_search:
        query = query.filter(ProductInstance.shelf_bin.ilike(f"%{bin_search}%"))

    # Always exclude sold items for inventory views
    if not status_filter or status_filter == 'all':
        query = query.filter(ProductInstance.is_sold == False)

    # Exclude all reserved units from inventory views
    reserved_order = aliased(CustomerOrderTracking)
    query = query.outerjoin(
        reserved_order,
        (reserved_order.product_instance_id == ProductInstance.id) &
        (reserved_order.status.ilike('reserved'))
    ).filter(reserved_order.id == None)

    # --- Tenant scoping: only show ProductInstances for current tenant ---
    query = query.filter(Product.tenant_id == current_user.tenant_id)
    # --------------------------------------------------------------------

    # Query all matching instances (no pagination since we're grouping)
    all_instances = query.all()
    # --- DEBUG: Print all grouped inventory ---

    # Group instances by (model + cpu)
    from collections import defaultdict
    grouped = defaultdict(list)
    for instance in all_instances:
        key = (instance.product.model if instance.product else "", instance.product.cpu if instance.product else "")
        grouped[key].append(instance)

    grouped_instances = []
    for key, instances in grouped.items():
        product_id = None
        for i in instances:
            if i.product and i.product.id:
                product_id = i.product.id
                break
        grouped_instances.append({
            "model": key[0],
            "cpu": key[1],
            "count": len(instances),
            "instances": instances,
            "product_id": product_id
        })

    # Sort grouped list
    sort_col = request.args.get('sort', 'model')
    sort_dir = request.args.get('sort_dir', 'asc')
    reverse  = (sort_dir == 'desc')
    if sort_col == 'count':
        grouped_instances.sort(key=lambda g: g['count'], reverse=reverse)
    elif sort_col == 'cpu':
        grouped_instances.sort(key=lambda g: (g['cpu'] or '').lower(), reverse=reverse)
    else:  # default: model
        grouped_instances.sort(key=lambda g: (g['model'] or '').lower(), reverse=reverse)

    all_models = sorted({i.product.model for i in all_instances if i.product and i.product.model})
    all_processors = sorted({i.product.cpu for i in all_instances if i.product and i.product.cpu})
    all_rams = sorted({i.product.ram for i in all_instances if i.product and i.product.ram})
    all_disks = sorted({i.product.disk1size for i in all_instances if i.product and i.product.disk1size})
    distinct_stages = db.session.query(ProductInstance.process_stage).distinct().all()
    distinct_teams = db.session.query(ProductInstance.team_assigned).distinct().all()
    # Dynamically derive unique locations from visible instances
    all_locations = sorted({i.location.name for i in all_instances if i.location and i.location.name})

    # --- Unified column structure logic ---
    from inventory_flask_app.models import TenantSettings
    # Load and correct column order
    tenant_settings = TenantSettings.query.filter_by(tenant_id=current_user.tenant_id).all()
    settings_dict = {s.key: s.value for s in tenant_settings}
    column_order = settings_dict.get("column_order_instance_table")

    # Unified structure
    default_columns = [
        "asset", "serial", "Item Name", "make", "model", "display", "cpu", "ram",
        "gpu1", "gpu2", "Grade", "location", "status", "process_stage",
        "team_assigned", "shelf_bin", "is_sold", "age_days", "asking_price",
        "vendor", "po_number", "label", "action"
    ]

    # Normalize column names per-item (prevents duplicates from old aliases + case variants)
    _alias_map = {
        # Old API names
        "model_number": "model",
        "product": "Item Name",
        "processor": "cpu",
        "video_card": "gpu1",
        "resolution": "display",
        "screen_size": "display",
        # Case/format variants saved by admin column reorder
        "item_name": "Item Name",
        "item name": "Item Name",
        "grade": "Grade",
        "team": "team_assigned",
        "gpu": "gpu1",
    }
    if column_order:
        seen = set()
        existing = []
        for col in column_order.split(","):
            col = _alias_map.get(col.strip(), col.strip())
            if col and col not in seen:
                seen.add(col)
                existing.append(col)
        for col in default_columns:
            if col not in seen:
                seen.add(col)
                existing.append(col)
        column_order = ",".join(existing)
    else:
        column_order = ",".join(default_columns)

    settings_dict["column_order_instance_table"] = column_order
    # Set new columns hidden by default so they don't appear unless the admin enables them
    for _nc in ('age_days', 'asking_price', 'vendor', 'po_number'):
        settings_dict.setdefault(f'show_column_{_nc}', 'false')
    # --- END Unified column structure logic ---

    from inventory_flask_app.models import CustomField
    list_custom_fields = CustomField.query.filter_by(
        tenant_id=current_user.tenant_id, show_in_list=True
    ).order_by(CustomField.sort_order).all()

    if request.args.get("partial") == "1":
        return render_template(
            "_instance_rows.html",
            grouped_instances=grouped_instances,
            settings=settings_dict,
            list_custom_fields=list_custom_fields
        )

    # Compute total unit count across all groups
    total_units = sum(len(g["instances"]) for g in grouped_instances)

    # Summary stats for the stats bar
    _status_counts: dict = {}
    for _inst in all_instances:
        _s = _inst.status or 'unprocessed'
        _status_counts[_s] = _status_counts.get(_s, 0) + 1
    summary_stats = {
        'groups': len(grouped_instances),
        'total': total_units,
        'unprocessed': _status_counts.get('unprocessed', 0),
        'under_process': _status_counts.get('under_process', 0),
        'processed': _status_counts.get('processed', 0),
        'idle': _status_counts.get('idle', 0),
        'disputed': _status_counts.get('disputed', 0),
    }

    # Location objects for bulk-move modal
    all_location_objs = Location.query.filter_by(tenant_id=current_user.tenant_id).order_by(Location.name).all()

    return render_template(
        "instance_table.html",
        grouped_instances=grouped_instances,
        models=sorted(all_models),
        processors=sorted(all_processors),
        rams=sorted(all_rams),
        disk1sizes=sorted(all_disks),
        stages=[s[0] for s in distinct_stages if s[0]],
        teams=[t[0] for t in distinct_teams if t[0]],
        selected_stage=stage_filter,
        selected_team=team_filter,
        title=(
            "Total Inventory" if not status_filter or status_filter == 'all'
            else "Unprocessed Inventory" if status_filter == 'unprocessed'
            else "Processed Inventory" if status_filter == 'processed'
            else "Under Process Inventory" if status_filter == 'under_process'
            else "Disputed Inventory" if status_filter == 'disputed'
            else "Inventory"
        ),
        locations=all_locations,
        location_objs=all_location_objs,
        settings=settings_dict,
        pagination=None,
        total_units=total_units,
        selected_location=location_filter,
        list_custom_fields=list_custom_fields,
        sort_col=sort_col,
        sort_dir=sort_dir,
        summary_stats=summary_stats,
    )


# --- Helper: build unit list and user list for the My Units tab ---
def _get_units_for_tab(user):
    """Returns (my_units, all_users, is_supervisor).
    Admins/supervisors see ALL assigned units across all techs; staff see only their own."""
    from inventory_flask_app.models import User as UserModel
    is_supervisor = user.role in ('admin', 'supervisor')
    if is_supervisor:
        units = (
            ProductInstance.query.join(Product)
            .filter(
                Product.tenant_id == user.tenant_id,
                ProductInstance.assigned_to_user_id != None,
                ProductInstance.status.in_(['under_process', 'processed'])
            )
            .order_by(ProductInstance.entered_stage_at.asc().nulls_last())
            .all()
        )
        users = UserModel.query.filter_by(tenant_id=user.tenant_id).order_by(UserModel.username).all()
    else:
        # Fix 3: join Product to enforce tenant isolation for staff
        units = (
            ProductInstance.query.join(Product)
            .filter(
                Product.tenant_id == user.tenant_id,
                ProductInstance.assigned_to_user_id == user.id,
                ProductInstance.status.in_(['under_process', 'processed'])
            )
            .all()
        )
        users = []
    return units, users, is_supervisor


# --- Tabbed Process Stage Management Route ---
@stock_bp.route('/process_stage/manage', methods=['GET'])
@login_required
def process_stage_update():
    """
    Tabbed view for process stage management, check-in, and check-out.
    All write actions go through /stock/checkin_checkout.
    """
    tab = request.args.get("tab") or "under_process"

    if tab == "under_process":
        my_units, all_users, is_supervisor = _get_units_for_tab(current_user)
        unprocessed_count = ProductInstance.query.join(Product).filter(
            Product.tenant_id == current_user.tenant_id,
            ProductInstance.status == 'unprocessed',
            ProductInstance.is_sold == False,
        ).count()
        return render_template(
            'process_stage_update.html',
            results=None,
            tab=tab,
            my_units=my_units,
            all_users=all_users,
            is_supervisor=is_supervisor,
            unprocessed_count=unprocessed_count,
        )
    elif tab == "check_in":
        # Restore locked stage/team from session if items already scanned (fix 13)
        scanned = session.get('scanned_checkin', [])
        locked_stage = scanned[0].get('stage', '') if scanned else ''
        locked_team = scanned[0].get('team', '') if scanned else ''
        # Pop check-in results from session for display (fixes dead results table)
        checkin_results = session.pop('checkin_results', None)
        session.modified = True
        # Load processing teams from TenantSettings
        from inventory_flask_app.models import TenantSettings
        _ts = TenantSettings.query.filter_by(
            tenant_id=current_user.tenant_id, key='processing_teams'
        ).first()
        _teams_raw = (_ts.value if _ts and _ts.value and _ts.value.strip() else
                      'Tech A,Tech B,Paint,QC')
        teams = [t.strip() for t in _teams_raw.split(',') if t.strip()]
        return render_template(
            'process_stage_update.html',
            results=checkin_results,
            tab=tab,
            locked_stage=locked_stage,
            locked_team=locked_team,
            teams=teams,
        )
    elif tab == "check_out":
        # Fix 2: pass my_units so the reference table renders
        my_units, all_users, is_supervisor = _get_units_for_tab(current_user)
        return render_template(
            'process_stage_update.html',
            results=None,
            tab=tab,
            my_units=my_units,
            all_users=all_users,
            is_supervisor=is_supervisor,
            unprocessed_count=0,
        )
    # Default: fallback to under_process tab
    my_units, all_users, is_supervisor = _get_units_for_tab(current_user)
    unprocessed_count = ProductInstance.query.join(Product).filter(
        Product.tenant_id == current_user.tenant_id,
        ProductInstance.status == 'unprocessed',
        ProductInstance.is_sold == False,
    ).count()
    return render_template(
        'process_stage_update.html',
        results=None,
        tab='under_process',
        my_units=my_units,
        all_users=all_users,
        is_supervisor=is_supervisor,
        unprocessed_count=unprocessed_count,
    )

@stock_bp.route('/instance/<int:instance_id>/view_edit', methods=['GET', 'POST'])
@login_required
def view_edit_instance(instance_id):
    # --- Tenant scoping: only allow view/edit of ProductInstances for current tenant ---
    instance = ProductInstance.query.join(Product).filter(
        ProductInstance.id == instance_id,
        Product.tenant_id == current_user.tenant_id
    ).first_or_404()
    # --------------------------------------------------------------------
    allowed_statuses = ['unprocessed', 'under_process', 'processed', 'sold', 'disputed']
    if request.method == 'POST':
        status = request.form.get('status', instance.status)
        if status not in allowed_statuses:
            flash("Invalid status selected.", "danger")
            return redirect(url_for('stock_bp.view_edit_instance', instance_id=instance.id))

        old_status = instance.status
        old_stage = instance.process_stage

        # Validate process_stage against configured stages
        from inventory_flask_app.models import ProcessStage
        _valid_stages = {s.name for s in ProcessStage.query.filter_by(tenant_id=current_user.tenant_id).all()}
        new_stage_raw = request.form.get('process_stage', '').strip() or None
        if new_stage_raw and _valid_stages and new_stage_raw not in _valid_stages:
            flash(f'Invalid stage "{new_stage_raw}". Choose a configured stage.', "danger")
            return redirect(url_for('stock_bp.view_edit_instance', instance_id=instance.id))

        instance.status = status
        instance.process_stage = new_stage_raw if new_stage_raw is not None else instance.process_stage
        instance.team_assigned = request.form.get('team_assigned', instance.team_assigned)
        # location and bin
        new_loc = request.form.get('location_id', '').strip()
        if new_loc:
            loc = Location.query.filter_by(id=int(new_loc), tenant_id=current_user.tenant_id).first()
            if loc:
                instance.location_id = loc.id
        new_bin = request.form.get('shelf_bin', '').strip().upper()
        if new_bin or request.form.get('shelf_bin') == '':
            instance.shelf_bin = new_bin or None
            # Resolve to structured Bin record if location is set
            if new_bin and instance.location_id:
                managed = Bin.query.filter_by(
                    name=new_bin, location_id=instance.location_id, tenant_id=current_user.tenant_id
                ).first()
                instance.bin_id = managed.id if managed else None
            elif not new_bin:
                instance.bin_id = None
        if status != 'under_process':
            instance.assigned_to_user_id = None
        if status == 'unprocessed':
            instance.process_stage = None

        # Time-tracking: log stage/status changes and update entered_stage_at
        now_ts = get_now_for_tenant()
        _final_stage = instance.process_stage
        if _final_stage != old_stage or status != old_status:
            duration = calc_duration_minutes(instance.entered_stage_at)
            db.session.add(ProductProcessLog(
                product_instance_id=instance.id,
                from_stage=old_stage,
                to_stage=_final_stage,
                moved_by=current_user.id,
                moved_at=now_ts,
                action='stage_change_manual',
                note=f'Manual edit by {current_user.username}: status {old_status}→{status}' if status != old_status else f'Manual stage edit by {current_user.username}',
                duration_minutes=duration,
            ))
            if _final_stage != old_stage:
                instance.entered_stage_at = now_ts if status == 'under_process' else None
            elif status not in ('under_process',):
                instance.entered_stage_at = None
            sync_reservation_stage(instance.id, _final_stage, current_user.username)

        # Save custom field values
        from inventory_flask_app.models import CustomField, CustomFieldValue
        custom_fields = CustomField.query.filter_by(tenant_id=current_user.tenant_id).order_by(CustomField.sort_order).all()
        for cf in custom_fields:
            form_key = f'cf_{cf.field_key}'
            raw_val = request.form.get(form_key)
            if raw_val is None:
                continue
            cfv = CustomFieldValue.query.filter_by(instance_id=instance.id, field_id=cf.id).first()
            if cfv:
                cfv.value = raw_val or None
            else:
                if raw_val:
                    db.session.add(CustomFieldValue(
                        tenant_id=current_user.tenant_id,
                        instance_id=instance.id,
                        field_id=cf.id,
                        value=raw_val
                    ))
        instance.asking_price = request.form.get('asking_price', type=float) or None
        db.session.commit()
        flash("Instance updated.", "success")
        return redirect(url_for('stock_bp.view_edit_instance', instance_id=instance.id))
    locations = Location.query.filter_by(tenant_id=current_user.tenant_id).order_by(Location.name).all()
    from inventory_flask_app.models import ProcessStage, CustomField, CustomFieldValue, Return
    try:
        process_stages = ProcessStage.query.filter_by(tenant_id=current_user.tenant_id).order_by(ProcessStage.order).all()
    except Exception:
        process_stages = []
    custom_fields = CustomField.query.filter_by(tenant_id=current_user.tenant_id).order_by(CustomField.sort_order).all()
    cf_values = {cfv.field_id: cfv.value for cfv in CustomFieldValue.query.filter_by(instance_id=instance.id).all()}
    instance_returns = Return.query.filter_by(
        instance_id=instance.id, tenant_id=current_user.tenant_id
    ).order_by(Return.return_date.desc()).all()
    from inventory_flask_app.models import CustomerOrderTracking
    active_reservation = CustomerOrderTracking.query.filter(
        CustomerOrderTracking.product_instance_id == instance.id,
        CustomerOrderTracking.status.in_(['reserved', 'delivered']),
    ).first()
    return render_template('view_edit_instance.html', instance=instance, locations=locations,
                           process_stages=process_stages, custom_fields=custom_fields, cf_values=cf_values,
                           instance_returns=instance_returns, active_reservation=active_reservation)


# ─────────────────────────────────────────────────────────────
# Supervisor: Reassign unit to another technician
# ─────────────────────────────────────────────────────────────
@stock_bp.route('/instance/<int:instance_id>/reassign', methods=['POST'])
@login_required
def reassign_instance(instance_id):
    from flask import abort
    from inventory_flask_app.models import User as UserModel
    if current_user.role not in ('admin', 'supervisor'):
        abort(403)

    instance = ProductInstance.query.join(Product).filter(
        ProductInstance.id == instance_id,
        Product.tenant_id == current_user.tenant_id
    ).first_or_404()

    new_user_id_raw = request.form.get('new_user_id', '').strip()
    if not new_user_id_raw:
        flash('No user selected.', 'danger')
        return redirect(url_for('stock_bp.process_stage_update', tab='under_process'))

    new_user = UserModel.query.filter_by(id=int(new_user_id_raw), tenant_id=current_user.tenant_id).first()
    if not new_user:
        flash('Invalid user.', 'danger')
        return redirect(url_for('stock_bp.process_stage_update', tab='under_process'))

    from_user = instance.assigned_user
    from_name = from_user.username if from_user else 'unassigned'
    duration = calc_duration_minutes(instance.entered_stage_at)
    now_ts = get_now_for_tenant()

    instance.assigned_to_user_id = new_user.id
    instance.entered_stage_at = now_ts  # reset clock for new assignee
    instance.updated_at = now_ts

    db.session.add(ProductProcessLog(
        product_instance_id=instance.id,
        from_stage=instance.process_stage,
        to_stage=instance.process_stage,
        from_team=instance.team_assigned,
        to_team=instance.team_assigned,
        moved_by=current_user.id,
        moved_at=now_ts,
        action='reassigned',
        note=f'Reassigned from {from_name} to {new_user.username} by {current_user.username}',
        duration_minutes=duration,
    ))
    # In-app notification to the newly assigned user
    create_notification(
        user_id=new_user.id,
        notif_type='reassigned',
        title='Unit Reassigned to You',
        message=f'{instance.serial} has been assigned to you by {current_user.username}',
        link=url_for('stock_bp.process_stage_update', tab='under_process'),
    )
    db.session.commit()
    flash(f'Unit {instance.serial} reassigned to {new_user.username}.', 'success')
    return redirect(url_for('stock_bp.process_stage_update', tab='under_process'))


# ─────────────────────────────────────────────────────────────
# Supervisor: Force check-out — release a stuck/locked unit
# ─────────────────────────────────────────────────────────────
@stock_bp.route('/instance/<int:instance_id>/force_checkout', methods=['POST'])
@login_required
def force_checkout(instance_id):
    from flask import abort
    if current_user.role not in ('admin', 'supervisor'):
        abort(403)

    instance = ProductInstance.query.join(Product).filter(
        ProductInstance.id == instance_id,
        Product.tenant_id == current_user.tenant_id
    ).first_or_404()

    from_user = instance.assigned_user
    from_name = from_user.username if from_user else 'unassigned'
    prev_stage = instance.process_stage
    prev_team = instance.team_assigned
    duration = calc_duration_minutes(instance.entered_stage_at)
    now_ts = get_now_for_tenant()

    instance.status = 'processed'
    instance.assigned_to_user_id = None
    instance.entered_stage_at = None
    instance.updated_at = now_ts

    db.session.add(ProductProcessLog(
        product_instance_id=instance.id,
        from_stage=prev_stage,
        to_stage=prev_stage,
        from_team=prev_team,
        to_team=prev_team,
        moved_by=current_user.id,
        moved_at=now_ts,
        action='force-checkout',
        note=f'Force checked out by {current_user.username} (was: {from_name})',
        duration_minutes=duration,
    ))
    db.session.commit()
    flash(f'Unit {instance.serial} force checked out from {from_name}.', 'success')
    return redirect(url_for('stock_bp.process_stage_update', tab='under_process'))


# ─────────────────────────────────────────────────────────────
# Return idle unit to unprocessed (any staff can do this)
# ─────────────────────────────────────────────────────────────
@stock_bp.route('/instance/<int:instance_id>/return_from_idle', methods=['POST'])
@login_required
def return_from_idle(instance_id):
    instance = ProductInstance.query.join(Product).filter(
        ProductInstance.id == instance_id,
        Product.tenant_id == current_user.tenant_id,
        ProductInstance.status == 'idle',
    ).first_or_404()

    now_ts = get_now_for_tenant()
    prev_reason = instance.idle_reason

    instance.status = 'unprocessed'
    instance.idle_reason = None
    instance.process_stage = None
    instance.assigned_to_user_id = None
    instance.entered_stage_at = None
    instance.updated_at = now_ts
    sync_reservation_stage(instance.id, None, current_user.username)

    db.session.add(ProductProcessLog(
        product_instance_id=instance.id,
        from_stage='idle',
        to_stage=None,
        moved_by=current_user.id,
        moved_at=now_ts,
        action='returned_from_idle',
        note=f'Returned to unprocessed by {current_user.username}' + (f' (was idle: {prev_reason})' if prev_reason else ''),
    ))
    db.session.commit()
    flash(f'Unit {instance.serial} returned to unprocessed.', 'success')
    return redirect(request.referrer or url_for('reports_bp.idle_units'))


# ─────────────────────────────────────────────────────────────
# Mark unit as disputed
# ─────────────────────────────────────────────────────────────
@stock_bp.route('/instance/<int:instance_id>/mark_disputed', methods=['POST'])
@login_required
def mark_disputed(instance_id):
    instance = ProductInstance.query.join(Product).filter(
        ProductInstance.id == instance_id,
        Product.tenant_id == current_user.tenant_id,
    ).first_or_404()

    reason = (request.form.get('reason') or '').strip()
    note = (request.form.get('note') or '').strip()
    now_ts = get_now_for_tenant()
    old_status = instance.status
    duration = calc_duration_minutes(instance.entered_stage_at)

    instance.status = 'disputed'
    instance.idle_reason = reason or None
    instance.updated_at = now_ts

    full_note = f'Marked disputed by {current_user.username}: {reason}' + (f' — {note}' if note else '')
    db.session.add(ProductProcessLog(
        product_instance_id=instance.id,
        from_stage=instance.process_stage,
        to_stage='disputed',
        moved_by=current_user.id,
        moved_at=now_ts,
        action='marked_disputed',
        note=full_note[:200],
        duration_minutes=duration,
    ))
    # Notify all supervisors and admins in this tenant
    from inventory_flask_app.models import User as _UserModel
    supervisors = _UserModel.query.filter(
        _UserModel.tenant_id == current_user.tenant_id,
        _UserModel.role.in_(('admin', 'supervisor')),
        _UserModel.id != current_user.id,
    ).all()
    for sup in supervisors:
        create_notification(
            user_id=sup.id,
            notif_type='disputed',
            title='Unit Marked Disputed',
            message=f'{instance.serial} marked disputed: {reason}',
            link=url_for('stock_bp.view_edit_instance', instance_id=instance_id),
        )
    db.session.commit()
    flash(f'Unit {instance.serial} marked as disputed.', 'warning')
    return redirect(request.referrer or url_for('stock_bp.view_edit_instance', instance_id=instance_id))


# ─────────────────────────────────────────────────────────────
# Resolve dispute (supervisor/admin only)
# ─────────────────────────────────────────────────────────────
@stock_bp.route('/instance/<int:instance_id>/resolve_dispute', methods=['POST'])
@login_required
def resolve_dispute(instance_id):
    from flask import abort
    if current_user.role not in ('admin', 'supervisor'):
        abort(403)

    instance = ProductInstance.query.join(Product).filter(
        ProductInstance.id == instance_id,
        Product.tenant_id == current_user.tenant_id,
    ).first_or_404()

    now_ts = get_now_for_tenant()
    instance.status = 'unprocessed'
    instance.idle_reason = None
    instance.process_stage = None
    instance.assigned_to_user_id = None
    instance.entered_stage_at = None
    instance.updated_at = now_ts
    sync_reservation_stage(instance.id, None, current_user.username)

    db.session.add(ProductProcessLog(
        product_instance_id=instance.id,
        from_stage='disputed',
        to_stage=None,
        moved_by=current_user.id,
        moved_at=now_ts,
        action='dispute_resolved',
        note=f'Dispute resolved by {current_user.username}',
    ))
    db.session.commit()
    flash(f'Dispute resolved for {instance.serial}. Unit returned to unprocessed.', 'success')
    return redirect(request.referrer or url_for('stock_bp.view_edit_instance', instance_id=instance_id))


# --- Unit history route ---
@stock_bp.route('/unit_history/<serial>')
@login_required
def unit_history(serial):
    from sqlalchemy import or_
    instance = ProductInstance.query.join(Product).filter(
        or_(
            ProductInstance.serial == serial,
            ProductInstance.asset == serial
        ),
        Product.tenant_id == current_user.tenant_id
    ).first_or_404()
    from inventory_flask_app.models import SaleTransaction, Customer
    sale = SaleTransaction.query.filter_by(product_instance_id=instance.id).join(Customer).first()
    logs = ProductProcessLog.query.filter_by(product_instance_id=instance.id).order_by(ProductProcessLog.moved_at).all()
    return render_template('unit_history.html', instance=instance, logs=logs, sale=sale)

# --- Read-only Unit Detail page ---
@stock_bp.route('/instance/<int:instance_id>/view')
@login_required
def unit_detail(instance_id):
    from inventory_flask_app.models import Return, SaleTransaction, Customer
    instance = ProductInstance.query.join(Product).filter(
        ProductInstance.id == instance_id,
        Product.tenant_id == current_user.tenant_id
    ).first_or_404()
    logs = (ProductProcessLog.query
            .filter_by(product_instance_id=instance.id)
            .order_by(ProductProcessLog.moved_at.desc())
            .all())
    instance_returns = (Return.query
                        .filter_by(instance_id=instance.id, tenant_id=current_user.tenant_id)
                        .order_by(Return.return_date.desc())
                        .all())
    sale = (SaleTransaction.query
            .filter_by(product_instance_id=instance.id)
            .join(Customer)
            .first())
    return render_template('stock/unit_detail.html',
                           instance=instance,
                           logs=logs,
                           instance_returns=instance_returns,
                           sale=sale)


# Delete ProductInstance route
@stock_bp.route('/instance/<int:instance_id>/delete', methods=['POST'])
@login_required
@admin_or_supervisor_required
def delete_instance(instance_id):
    # --- Tenant scoping: only allow delete of ProductInstances for current tenant ---
    instance = ProductInstance.query.join(Product).filter(
        ProductInstance.id == instance_id,
        Product.tenant_id == current_user.tenant_id
    ).first_or_404()
    # --------------------------------------------------------------------
    db.session.delete(instance)
    db.session.commit()
    flash(f"✅ Unit with serial '{instance.serial}' deleted successfully.", "success")
    return redirect(url_for('stock_bp.under_process'))

# --- Improved Check-in/Check-out route with assignment logic ---
@stock_bp.route('/checkin_checkout', methods=['POST'])
@login_required
def checkin_checkout():
    # Handles technician check-in / check-out actions from Processing page.
    action = (request.form.get("action") or "").strip()  # "check-in" or "check-out"
    tab = request.form.get("tab") or request.args.get("tab")

    # IDs posted from the form (preferred)
    instance_ids = [s for s in request.form.getlist("instance_ids") if s]

    # Optional: mark idle ids (from Under Process tab)
    idle_ids = set([s for s in request.form.getlist("mark_idle_ids") if s])

    # Notes payload (may include notes[<id>] or notes[<serial>])
    notes_dict = request.form.to_dict(flat=False)

    # Fallback stage/team (used for check-in if per-scan values missing)
    fallback_stage = (request.form.get("process_stage") or "").strip()
    fallback_team = (request.form.get("team_assigned") or "").strip()

    scanned_checkin = session.get("scanned_checkin", []) or []
    scanned_checkout = session.get("scanned_checkout", []) or []

    # If no IDs were posted, fall back to session scanned lists
    if not instance_ids:
        if action == "check-in" and scanned_checkin:
            scanned_serials = [str(e.get("serial", "")).strip() for e in scanned_checkin if e.get("serial")]
            if scanned_serials:
                matched = ProductInstance.query.join(Product).filter(
                    ProductInstance.serial.in_(scanned_serials),
                    Product.tenant_id == current_user.tenant_id
                ).all()
                instance_ids = [str(i.id) for i in matched]
                # Inject notes from session into notes_dict as notes[<id>]
                for inst in matched:
                    entry = next((e for e in scanned_checkin if str(e.get("serial", "")).strip() == inst.serial), None)
                    if entry and entry.get("note"):
                        notes_dict[f"notes[{inst.id}]"] = [entry["note"]]
        elif action == "check-out" and scanned_checkout:
            scanned_serials = [str(e.get("serial", "")).strip() for e in scanned_checkout if e.get("serial")]
            if scanned_serials:
                matched = ProductInstance.query.join(Product).filter(
                    ProductInstance.serial.in_(scanned_serials),
                    Product.tenant_id == current_user.tenant_id
                ).all()
                instance_ids = [str(i.id) for i in matched]
                for inst in matched:
                    entry = next((e for e in scanned_checkout if str(e.get("serial", "")).strip() == inst.serial), None)
                    if entry and entry.get("note"):
                        notes_dict[f"notes[{inst.id}]"] = [entry["note"]]

    # If still nothing to process and no idle ids, stop
    if not instance_ids and not idle_ids:
        flash("Missing information for check-in/out.", "danger")
        if not tab:
            tab = "check_in" if action == "check-in" else "check_out"
        return redirect(url_for('stock_bp.process_stage_update', tab=tab))

    updated_count = 0
    skipped = 0
    checkin_results = []  # track per-unit results for check-in summary table

    # Process union of instance ids and idle ids
    all_ids = set(instance_ids).union(idle_ids)

    for iid in all_ids:
        try:
            iid_int = int(iid)
        except Exception:
            continue

        instance = ProductInstance.query.join(Product).filter(
            ProductInstance.id == iid_int,
            Product.tenant_id == current_user.tenant_id
        ).first()
        if not instance:
            continue

        # Note lookup (notes[<id>]) with fallback to notes[<serial>]
        note_key = f"notes[{iid_int}]"
        note = ""
        if note_key in notes_dict and notes_dict[note_key]:
            note = notes_dict[note_key][0] or ""
        else:
            serial_key = f"notes[{instance.serial}]"
            if serial_key in notes_dict and notes_dict[serial_key]:
                note = notes_dict[serial_key][0] or ""

        # Mark idle (if selected)
        if str(iid_int) in idle_ids:
            prev_stage = instance.process_stage
            prev_team = instance.team_assigned
            duration = calc_duration_minutes(instance.entered_stage_at)
            now_ts = get_now_for_tenant()
            instance.status = "idle"
            instance.idle_reason = note or None  # fix 6: populate idle_reason
            instance.assigned_to_user_id = None
            instance.entered_stage_at = None
            instance.updated_at = now_ts
            db.session.add(ProductProcessLog(
                product_instance_id=instance.id,
                from_stage=prev_stage,
                to_stage="idle",
                from_team=prev_team,
                to_team=prev_team,
                moved_by=current_user.id,
                moved_at=now_ts,
                action="moved_to_idle",
                note=note,
                duration_minutes=duration,
            ))
            updated_count += 1
            continue

        if action == "check-in":
            # Capture product info for results (before any changes)
            _prod = instance.product
            _result_base = {
                'serial': instance.serial, 'asset': instance.asset or '',
                'instance_id': instance.id,
                'item_name': _prod.item_name if _prod else '',
                'make': _prod.make if _prod else '',
                'model': _prod.model if _prod else '',
                'cpu': _prod.cpu if _prod else '',
                'ram': _prod.ram if _prod else '',
                'grade': _prod.grade if _prod else '',
                'display': _prod.display if _prod else '',
                'gpu1': _prod.gpu1 if _prod else '',
                'gpu2': _prod.gpu2 if _prod else '',
                'disk1size': _prod.disk1size if _prod else '',
                'prev_stage': instance.process_stage,
            }

            # Prevent check-in if assigned to someone else
            if instance.assigned_to_user_id and instance.assigned_to_user_id != current_user.id:
                skipped += 1
                checkin_results.append({**_result_base, 'status': 'skipped'})
                continue

            # Apply stage/team: prefer per-scan values if present
            stage_to_apply = fallback_stage
            team_to_apply = fallback_team
            if scanned_checkin:
                entry = next((e for e in scanned_checkin if str(e.get("serial", "")).strip().lower() == (instance.serial or '').lower()), None)
                if entry:
                    stage_to_apply = (entry.get("stage") or stage_to_apply).strip()
                    team_to_apply = (entry.get("team") or team_to_apply).strip()

            if not stage_to_apply or not team_to_apply:
                skipped += 1
                checkin_results.append({**_result_base, 'status': 'skipped'})
                continue

            prev_stage = instance.process_stage
            prev_team = instance.team_assigned

            # Time tracking: calculate how long unit was in previous stage
            duration = calc_duration_minutes(instance.entered_stage_at) if prev_stage else None

            now_ts = get_now_for_tenant()
            instance.status = "under_process"
            instance.process_stage = stage_to_apply
            instance.team_assigned = team_to_apply
            instance.assigned_to_user_id = current_user.id
            instance.updated_at = now_ts
            instance.entered_stage_at = now_ts  # clock starts for new stage
            sync_reservation_stage(instance.id, stage_to_apply, current_user.username)

            db.session.add(ProductProcessLog(
                product_instance_id=instance.id,
                from_stage=prev_stage,
                to_stage=stage_to_apply,
                from_team=prev_team,
                to_team=team_to_apply,
                moved_by=current_user.id,
                moved_at=now_ts,
                action="check-in",
                note=note,
                duration_minutes=duration,
            ))
            checkin_results.append({**_result_base, 'status': 'updated'})
            updated_count += 1

        elif action == "check-out":
            # Only allow check-out if assigned to current user
            if instance.assigned_to_user_id != current_user.id:
                skipped += 1
                continue

            prev_stage = instance.process_stage
            prev_team = instance.team_assigned

            # Time tracking: total time in the final stage
            duration = calc_duration_minutes(instance.entered_stage_at)

            now_ts = get_now_for_tenant()
            instance.status = "processed"
            instance.process_stage = None  # fix 8: clear stage on checkout
            instance.assigned_to_user_id = None
            instance.entered_stage_at = None  # clear on checkout
            instance.updated_at = now_ts
            sync_reservation_stage(instance.id, None, current_user.username)

            db.session.add(ProductProcessLog(
                product_instance_id=instance.id,
                from_stage=prev_stage,
                to_stage='processed',
                from_team=prev_team,
                to_team=prev_team,
                moved_by=current_user.id,
                moved_at=now_ts,
                action="check-out",
                note=note,
                duration_minutes=duration,
            ))
            updated_count += 1

        else:
            flash("Missing information for check-in/out.", "danger")
            if not tab:
                tab = "check_in"
            return redirect(url_for('stock_bp.process_stage_update', tab=tab))

    db.session.commit()

    # Clear session scan lists after processing; store check-in results for display
    if action == "check-in":
        session.pop("scanned_checkin", None)
        if checkin_results:
            session['checkin_results'] = checkin_results
    if action == "check-out":
        session.pop("scanned_checkout", None)
    session.modified = True

    msg = f"✅ {action} complete: {updated_count} unit(s) updated."
    if skipped:
        # fix 14: explain WHY units were skipped
        if action == "check-in":
            msg += f" {skipped} skipped — either stage/team not locked or unit assigned to another technician."
        else:
            msg += f" {skipped} skipped — unit not assigned to you."
    flash(msg, "success" if updated_count else "warning")

    if not tab:
        tab = "check_in" if action == "check-in" else "check_out"
    return redirect(url_for('stock_bp.process_stage_update', tab=tab))

@stock_bp.route('/export_instances', methods=['GET', 'POST'])
@login_required
def export_instances():
    from openpyxl import Workbook
    from io import BytesIO
    from flask import send_file
    from sqlalchemy import or_

    status_filter = request.args.get('status')
    model_filter = request.args.get('model')
    processor_filter = request.args.get('processor')
    serial_search = request.args.get('serial_search', '').strip()
    stage_filter = request.args.get('stage')
    team_filter = request.args.get('team')
    # location_id = request.args.get('location_id')
    location_filter = request.args.get('location', '').strip()
    bin_search = request.args.get('bin_search', '').strip()
    ram_filter = request.args.get('ram')
    disk_filter = request.args.get('disk1size')

    # --- Export selected logic ---
    instance_ids = request.form.getlist('instance_ids') if request.method == 'POST' else []

    # --- Expand grouped keys (model|||cpu) if present ---
    from sqlalchemy import or_, and_
    grouped_keys = [v for v in instance_ids if "|||" in v]
    exact_ids = [int(v) for v in instance_ids if v.isdigit()]

    if not status_filter or status_filter == 'all':
        query = ProductInstance.query
    else:
        query = ProductInstance.query
        if status_filter == 'unprocessed':
            query = query.filter(ProductInstance.status == 'unprocessed', ProductInstance.is_sold == False)
        elif status_filter == 'under_process':
            query = query.filter(ProductInstance.status == 'under_process', ProductInstance.is_sold == False)
        elif status_filter == 'processed':
            query = query.filter(ProductInstance.status == 'processed', ProductInstance.is_sold == False)
        elif status_filter == 'disputed':
            query = query.filter(ProductInstance.status == 'disputed', ProductInstance.is_sold == False)
        else:
            query = query.filter(ProductInstance.status == status_filter)

    # Eager load related product and location data for performance
    query = query.options(
        joinedload(ProductInstance.product),
        joinedload(ProductInstance.location)
    )

    # Join Product only once if any filter is present (or always, to simplify filter logic)
    if model_filter or processor_filter or ram_filter or disk_filter or True:
        query = query.join(Product)
    if model_filter:
        query = query.filter(Product.model.ilike(f"%{model_filter}%"))
    if processor_filter:
        query = query.filter(Product.cpu == processor_filter)
    # RAM and Disk filter logic (re-added after model/processor filter)
    if ram_filter:
        query = query.filter(Product.ram == ram_filter)
    if disk_filter:
        query = query.filter(Product.disk1size == disk_filter)
    if serial_search:
        query = query.filter(or_(
            ProductInstance.serial.ilike(f"%{serial_search}%"),
            ProductInstance.asset.ilike(f"%{serial_search}%")
        ))
    if stage_filter:
        query = query.filter(ProductInstance.process_stage == stage_filter)
    if team_filter:
        query = query.filter(ProductInstance.team_assigned.ilike(f"%{team_filter}%"))
    # Location name filter (new)
    if location_filter:
        query = query.join(ProductInstance.location).filter(Location.name.ilike(location_filter))
    # If you want to keep supporting location_id filter (by id), you may add:
    # if location_id:
    #     query = query.filter(ProductInstance.location_id == int(location_id))
    if bin_search:
        query = query.filter(ProductInstance.shelf_bin.ilike(f"%{bin_search}%"))

    # Always exclude sold items for inventory views
    if not status_filter or status_filter == 'all':
        query = query.filter(ProductInstance.is_sold == False)

    # Exclude all reserved units from inventory views
    reserved_order = aliased(CustomerOrderTracking)
    query = query.outerjoin(
        reserved_order,
        (reserved_order.product_instance_id == ProductInstance.id) &
        (reserved_order.status.ilike('reserved'))
    ).filter(reserved_order.id == None)

    # --- Tenant scoping: only show ProductInstances for current tenant ---
    query = query.filter(Product.tenant_id == current_user.tenant_id)
    # --------------------------------------------------------------------

    # --- Export selected logic: restrict to selected IDs or grouped keys if provided ---
    if grouped_keys:
        filters = []
        for key in grouped_keys:
            m, c = key.split("|||")
            filters.append(and_(Product.model == m.strip(), Product.cpu == c.strip()))
        model_cpu_filter = or_(*filters)
        query = query.filter(model_cpu_filter)

    if exact_ids:
        query = query.filter(ProductInstance.id.in_(exact_ids))
    # -------------------------------------------------------------------

    instances = query.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory Export"
    ws.append([
        "Serial", "Asset", "Item Name", "Make", "Model", "CPU", "RAM", "Disk", "Display",
        "GPU1", "GPU2", "Grade", "Location", "Status", "Process Stage", "Shelf Bin", "Team", "Sold"
    ])

    for i in instances:
        p = i.product
        ws.append([
            i.serial,
            i.asset,
            p.item_name if p else '',
            p.make if p else '',
            p.model if p else '',
            p.cpu if p else '',
            p.ram if p else '',
            p.disk1size if p else '',
            p.display if p else '',
            p.gpu1 if p else '',
            p.gpu2 if p else '',
            p.grade if p else '',
            i.location.name if i.location else '',
            i.status,
            i.process_stage or '',
            i.shelf_bin or '',
            i.team_assigned or '',
            'Yes' if i.is_sold else 'No'
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name="inventory_export.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# scan_move moved to routes/scanner.py (scanner_bp)


# QR code label printing route
@stock_bp.route('/print_label/<int:instance_id>')
@login_required
def print_label(instance_id):
    from datetime import datetime
    # Tenant-scoped query for ProductInstance
    instance = ProductInstance.query.join(Product).filter(
        ProductInstance.id == instance_id,
        Product.tenant_id == current_user.tenant_id
    ).first_or_404()
    if not instance.product:
        flash("No product linked to this instance.", "danger")
        return redirect(request.referrer or url_for('dashboard_bp.main_dashboard'))

    # Encode key info as QR code using unified product structure
    qr_data = {
        "serial": instance.serial if instance else "",
        "asset": instance.asset if instance else "",
        "item_name": instance.product.item_name if instance.product else "",
        "make": instance.product.make if instance.product else "",
        "model": instance.product.model if instance.product else "",
        "display": instance.product.display if instance.product else "",
        "cpu": instance.product.cpu if instance.product else "",
        "ram": instance.product.ram if instance.product else "",
        "gpu1": instance.product.gpu1 if instance.product else "",
        "gpu2": instance.product.gpu2 if instance.product else "",
        "grade": instance.product.grade if instance.product else ""
    }
    qr = qrcode.QRCode(version=1, box_size=10, border=2)
    qr.add_data(qr_data)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")

    # Save QR code image to base64 for embedding in HTML
    buffered = BytesIO()
    img.save(buffered, format="PNG")
    qr_b64 = base64.b64encode(buffered.getvalue()).decode()

    return render_template(
        "print_label.html",
        instance=instance,
        qr_b64=qr_b64,
        printed_time=datetime.now()
    )

@stock_bp.route('/batch_update_status', methods=['POST'])
@login_required
def batch_update_status():
    ids = request.form.getlist('instance_ids')
    status = request.form.get('status')
    location_id = request.form.get('location_id')
    if not ids or not status:
        flash("Please select at least one item and a status.", "danger")
        return redirect(request.referrer or url_for('dashboard_bp.main_dashboard'))

    # Only allow certain statuses for safety
    allowed_statuses = ['unprocessed', 'under_process', 'processed', 'sold']
    if status not in allowed_statuses:
        flash("Invalid status selected.", "danger")
        return redirect(request.referrer or url_for('dashboard_bp.main_dashboard'))

    # Enforce tenant scoping: only update ProductInstances for current tenant
    updated_count = 0
    # Join Product to ensure tenant filtering
    for instance in ProductInstance.query.join(Product).filter(
        ProductInstance.id.in_(ids),
        Product.tenant_id == current_user.tenant_id
    ).all():
        instance.status = status
        if location_id:
            instance.location_id = int(location_id)
        updated_count += 1
    db.session.commit()

    flash(f"✅ {updated_count} item(s) updated to '{status.replace('_', ' ').title()}' status.", "success")
    return redirect(request.referrer or url_for('dashboard_bp.main_dashboard'))

# Add product page route
@stock_bp.route('/products/add', methods=['GET', 'POST'])
@login_required
def add_product_page():
    from inventory_flask_app.models import TenantSettings
    tenant_settings = TenantSettings.query.filter_by(tenant_id=current_user.tenant_id).all()
    settings = {s.key: s.value for s in tenant_settings}

    if request.method == 'POST':
        data = request.form.to_dict()
        serial = (data.get("serial") or "").strip()
        if not serial:
            flash("Serial number is required.", "danger")
            return render_template('add_product.html', settings=settings)

        product = Product(
            item_name=data.get("item_name", "").strip(),
            model=data.get("model", "").strip(),
            cpu=data.get("cpu", "").strip(),
            ram=data.get("ram", "").strip(),
            disk1size=data.get("disk1size", "").strip(),
            display=data.get("display", "").strip(),
            gpu1=data.get("gpu1", "").strip(),
            gpu2=data.get("gpu2", "").strip(),
            grade=data.get("grade", "").strip(),
            tenant_id=current_user.tenant_id
        )
        db.session.add(product)
        db.session.flush()

        instance = ProductInstance(
            serial=serial,
            status=data.get("status", "unprocessed"),
            product_id=product.id,
            tenant_id=current_user.tenant_id
        )
        db.session.add(instance)
        db.session.commit()
        flash("✅ Product and instance added successfully!", "success")
        return redirect(url_for('stock_bp.print_label', instance_id=instance.id))

    return render_template('add_product.html', settings=settings)

@stock_bp.route('/bin_lookup', methods=['GET', 'POST'])
@login_required
def bin_lookup():
    if request.method == 'POST':
        bin_code = request.form.get('bin_code', '').strip().upper()
        if bin_code:
            # Check if a Bin record exists for this code; if so, use the clean /stock/bin/<id> route
            managed = Bin.query.filter_by(name=bin_code, tenant_id=current_user.tenant_id).first()
            if managed:
                return redirect(url_for('stock_bp.bin_detail', bin_id=managed.id))
            # Multiple bins with same name in different locations → show picker via legacy route
            return redirect(url_for('stock_bp.bin_contents', bin_code=bin_code))
    # Fetch locations with their bin counts for the browse panel
    locations = Location.query.filter_by(tenant_id=current_user.tenant_id).order_by(Location.name).all()
    loc_bins = {loc.id: Bin.query.filter_by(location_id=loc.id, tenant_id=current_user.tenant_id)
                                  .order_by(Bin.name).all()
                for loc in locations}
    return render_template('bin_lookup.html', locations=locations, loc_bins=loc_bins)

@stock_bp.route('/bin_contents/<bin_code>')
@login_required
def bin_contents(bin_code):
    bin_code = bin_code.upper()
    location_id = request.args.get('location_id', type=int)
    bin_id = request.args.get('bin_id', type=int)

    # Resolve by bin_id first (structured Bin record), fall back to shelf_bin text match
    managed_bin = None
    if bin_id:
        managed_bin = Bin.query.filter_by(id=bin_id, tenant_id=current_user.tenant_id).first()
    if not managed_bin and location_id:
        managed_bin = Bin.query.filter_by(
            name=bin_code, location_id=location_id, tenant_id=current_user.tenant_id
        ).first()

    if managed_bin:
        # Use structured FK lookup
        products = ProductInstance.query.join(Product).filter(
            ProductInstance.bin_id == managed_bin.id,
            ProductInstance.is_sold == False,
            Product.tenant_id == current_user.tenant_id
        ).all()
        # Also include legacy shelf_bin instances not yet linked to a Bin record
        legacy = ProductInstance.query.join(Product).filter(
            ProductInstance.bin_id == None,
            ProductInstance.shelf_bin == managed_bin.name,
            ProductInstance.location_id == managed_bin.location_id,
            ProductInstance.is_sold == False,
            Product.tenant_id == current_user.tenant_id
        ).all()
        products = products + legacy
        return render_template('bin_contents.html', bin_code=bin_code, products=products,
                               location=managed_bin.location, managed_bin=managed_bin)

    # Legacy: text-based shelf_bin lookup
    base_q = ProductInstance.query.join(Product).filter(
        ProductInstance.shelf_bin == bin_code,
        ProductInstance.is_sold == False,
        Product.tenant_id == current_user.tenant_id
    )

    if location_id:
        location = Location.query.filter_by(id=location_id, tenant_id=current_user.tenant_id).first_or_404()
        products = base_q.filter(ProductInstance.location_id == location_id).all()
        return render_template('bin_contents.html', bin_code=bin_code, products=products, location=location)

    from sqlalchemy import distinct
    loc_ids = [
        row[0] for row in
        ProductInstance.query.join(Product).filter(
            ProductInstance.shelf_bin == bin_code,
            ProductInstance.is_sold == False,
            Product.tenant_id == current_user.tenant_id,
            ProductInstance.location_id != None
        ).with_entities(distinct(ProductInstance.location_id)).all()
    ]

    if len(loc_ids) > 1:
        locations_with_bin = Location.query.filter(
            Location.id.in_(loc_ids),
            Location.tenant_id == current_user.tenant_id
        ).order_by(Location.name).all()
        return render_template('bin_contents.html', bin_code=bin_code, products=None,
                               location_picker=locations_with_bin)

    products = base_q.all()
    location = None
    if loc_ids:
        location = db.session.get(Location, loc_ids[0])
    return render_template('bin_contents.html', bin_code=bin_code, products=products, location=location)
# --- Add Location Route ---
# --- Add Location Route ---
@stock_bp.route('/location/add', methods=['GET', 'POST'])
@login_required
def add_location():
    from flask_wtf.csrf import validate_csrf
    from wtforms.validators import ValidationError

    if request.method == 'GET':
        locations = Location.query.filter_by(tenant_id=current_user.tenant_id).all()
        next_url = request.args.get('next') or request.referrer or url_for('stock_bp.manage_locations')
        return render_template("add_location.html", locations=locations, next_url=next_url)

    next_url = request.args.get('next') or request.form.get('next') or url_for('stock_bp.manage_locations')

    try:
        validate_csrf(request.form.get('csrf_token'))
    except ValidationError:
        flash("Invalid or missing CSRF token.", "danger")
        return redirect(next_url)

    name = request.form.get('name', '').strip().upper()
    if not name:
        flash("Location name is required.", "warning")
        return redirect(next_url)

    existing = Location.query.filter_by(name=name, tenant_id=current_user.tenant_id).first()
    if existing:
        flash("A location with this name already exists.", "warning")
        return redirect(next_url)

    location = Location(name=name, tenant_id=current_user.tenant_id)
    db.session.add(location)
    db.session.commit()
    flash(f"Location '{name}' added successfully.", "success")
    return redirect(next_url)


# --- Location Management ---
@stock_bp.route('/locations')
@login_required
def manage_locations():
    locations = Location.query.filter_by(
        tenant_id=current_user.tenant_id
    ).order_by(Location.name).all()

    loc_ids = [l.id for l in locations]

    # Unit counts + status breakdown per location
    unit_counts = {}
    status_breakdown = {}
    if loc_ids:
        for row in db.session.query(
            ProductInstance.location_id,
            func.count(ProductInstance.id).label('total'),
            func.sum(db.case((ProductInstance.status == 'processed', 1), else_=0)).label('processed'),
            func.sum(db.case((ProductInstance.status == 'under_process', 1), else_=0)).label('under_process'),
            func.sum(db.case((ProductInstance.status == 'unprocessed', 1), else_=0)).label('unprocessed'),
        ).join(Product).filter(
            Product.tenant_id == current_user.tenant_id,
            ProductInstance.is_sold == False,
            ProductInstance.location_id.in_(loc_ids),
        ).group_by(ProductInstance.location_id).all():
            unit_counts[row.location_id] = row.total
            status_breakdown[row.location_id] = {
                'processed': row.processed or 0,
                'under_process': row.under_process or 0,
                'unprocessed': row.unprocessed or 0,
            }

    # Bin counts per location split by type
    unit_bin_counts = {}
    parts_bin_counts = {}
    if loc_ids:
        for row in db.session.query(
            Bin.location_id,
            Bin.bin_type,
            func.count(Bin.id)
        ).filter(
            Bin.location_id.in_(loc_ids),
            Bin.tenant_id == current_user.tenant_id,
        ).group_by(Bin.location_id, Bin.bin_type).all():
            if row[1] == 'parts':
                parts_bin_counts[row[0]] = row[2]
            else:
                unit_bin_counts[row[0]] = row[2]

    return render_template(
        'locations.html',
        locations=locations,
        counts=unit_counts,
        unit_bin_counts=unit_bin_counts,
        parts_bin_counts=parts_bin_counts,
        status_breakdown=status_breakdown,
    )


# --- AJAX: create location on-the-fly ---
@stock_bp.route('/locations/create', methods=['POST'])
@login_required
def create_location_ajax():
    data = request.get_json() or {}
    name = (data.get('name') or '').strip().upper()
    if not name:
        return jsonify({'error': 'Name is required'}), 400
    existing = Location.query.filter_by(
        name=name, tenant_id=current_user.tenant_id
    ).first()
    if existing:
        return jsonify({'error': f"Location '{name}' already exists"}), 409
    loc = Location(name=name, tenant_id=current_user.tenant_id)
    db.session.add(loc)
    db.session.commit()
    return jsonify({'id': loc.id, 'name': loc.name})


@stock_bp.route('/locations/<int:location_id>/rename', methods=['POST'])
@login_required
def rename_location(location_id):
    location = Location.query.filter_by(id=location_id, tenant_id=current_user.tenant_id).first_or_404()
    new_name = request.form.get('name', '').strip().upper()
    if not new_name:
        flash("Location name is required.", "warning")
        return redirect(url_for('stock_bp.manage_locations'))
    existing = Location.query.filter_by(name=new_name, tenant_id=current_user.tenant_id).filter(Location.id != location_id).first()
    if existing:
        flash(f"A location named '{new_name}' already exists.", "warning")
        return redirect(url_for('stock_bp.manage_locations'))
    location.name = new_name
    db.session.commit()
    flash(f"Location renamed to '{new_name}'.", "success")
    return redirect(url_for('stock_bp.manage_locations'))


@stock_bp.route('/locations/<int:location_id>/delete', methods=['POST'])
@login_required
def delete_location(location_id):
    if current_user.role not in ('admin', 'supervisor'):
        flash("Admin access required.", "danger")
        return redirect(url_for('stock_bp.manage_locations'))
    location = Location.query.filter_by(id=location_id, tenant_id=current_user.tenant_id).first_or_404()
    unit_count = ProductInstance.query.join(Product).filter(
        ProductInstance.location_id == location_id,
        ProductInstance.is_sold == False,
        Product.tenant_id == current_user.tenant_id
    ).count()
    if unit_count > 0:
        flash(f"Cannot delete '{location.name}' — {unit_count} active unit(s) assigned.", "danger")
        return redirect(url_for('stock_bp.manage_locations'))
    db.session.delete(location)
    db.session.commit()
    flash(f"Location '{location.name}' deleted.", "success")
    return redirect(url_for('stock_bp.manage_locations'))


# ── Bin Management ─────────────────────────────────────────────────────────────

@stock_bp.route('/location/<int:location_id>/bins')
@login_required
def manage_bins(location_id):
    location = Location.query.filter_by(id=location_id, tenant_id=current_user.tenant_id).first_or_404()
    bins = Bin.query.filter_by(location_id=location_id, tenant_id=current_user.tenant_id).order_by(Bin.name).all()
    bin_ids = [b.id for b in bins]
    # Count units per bin
    counts = {}
    part_counts = {}
    if bin_ids:
        for row in db.session.query(
            ProductInstance.bin_id,
            func.count(ProductInstance.id)
        ).join(Product).filter(
            ProductInstance.bin_id.in_(bin_ids),
            ProductInstance.is_sold == False,
            Product.tenant_id == current_user.tenant_id
        ).group_by(ProductInstance.bin_id).all():
            counts[row[0]] = row[1]
        for row in db.session.query(
            PartStock.bin_id,
            func.count(PartStock.id)
        ).join(Part).filter(
            PartStock.bin_id.in_(bin_ids),
            Part.tenant_id == current_user.tenant_id,
        ).group_by(PartStock.bin_id).all():
            part_counts[row[0]] = row[1]
    return render_template('bins.html', location=location, bins=bins, counts=counts, part_counts=part_counts)


@stock_bp.route('/location/<int:location_id>/bins/add', methods=['POST'])
@login_required
def add_bin(location_id):
    location = Location.query.filter_by(id=location_id, tenant_id=current_user.tenant_id).first_or_404()
    name = request.form.get('name', '').strip().upper()
    description = request.form.get('description', '').strip() or None
    if not name:
        flash("Bin name is required.", "warning")
        return redirect(url_for('stock_bp.manage_bins', location_id=location_id))
    existing = Bin.query.filter_by(name=name, location_id=location_id, tenant_id=current_user.tenant_id).first()
    if existing:
        flash(f"Bin '{name}' already exists in {location.name}.", "warning")
        return redirect(url_for('stock_bp.manage_bins', location_id=location_id))
    bin_type = request.form.get('bin_type', 'units')
    if bin_type not in ('units', 'parts'):
        bin_type = 'units'
    from inventory_flask_app.utils import get_now_for_tenant
    b = Bin(name=name, location_id=location_id, tenant_id=current_user.tenant_id,
            description=description, bin_type=bin_type, created_at=get_now_for_tenant())
    db.session.add(b)
    db.session.commit()
    flash(f"Bin '{name}' added to {location.name}.", "success")
    return redirect(url_for('stock_bp.manage_bins', location_id=location_id))


@stock_bp.route('/location/<int:location_id>/bins/<int:bin_id>/edit', methods=['POST'])
@login_required
def edit_bin(location_id, bin_id):
    b = Bin.query.filter_by(id=bin_id, location_id=location_id, tenant_id=current_user.tenant_id).first_or_404()
    new_name = request.form.get('name', '').strip().upper()
    description = request.form.get('description', '').strip() or None
    if not new_name:
        flash("Bin name is required.", "warning")
        return redirect(url_for('stock_bp.manage_bins', location_id=location_id))
    conflict = Bin.query.filter(
        Bin.name == new_name,
        Bin.location_id == location_id,
        Bin.tenant_id == current_user.tenant_id,
        Bin.id != bin_id
    ).first()
    if conflict:
        flash(f"A bin named '{new_name}' already exists in this location.", "warning")
        return redirect(url_for('stock_bp.manage_bins', location_id=location_id))
    new_type = request.form.get('bin_type', b.bin_type)
    if new_type not in ('units', 'parts'):
        new_type = b.bin_type
    old_name = b.name
    b.name = new_name
    b.description = description
    b.bin_type = new_type
    # Keep shelf_bin in sync on linked instances
    if old_name != new_name:
        ProductInstance.query.filter_by(
            bin_id=bin_id, tenant_id=current_user.tenant_id
        ).update({'shelf_bin': new_name})
    db.session.commit()
    flash(f"Bin renamed to '{new_name}'.", "success")
    return redirect(url_for('stock_bp.manage_bins', location_id=location_id))


@stock_bp.route('/location/<int:location_id>/bins/<int:bin_id>/delete', methods=['POST'])
@login_required
def delete_bin(location_id, bin_id):
    if current_user.role not in ('admin', 'supervisor'):
        flash("Admin access required.", "danger")
        return redirect(url_for('stock_bp.manage_bins', location_id=location_id))
    b = Bin.query.filter_by(id=bin_id, location_id=location_id, tenant_id=current_user.tenant_id).first_or_404()
    unit_count = ProductInstance.query.join(Product).filter(
        ProductInstance.bin_id == bin_id,
        ProductInstance.is_sold == False,
        Product.tenant_id == current_user.tenant_id
    ).count()
    if unit_count > 0:
        flash(f"Cannot delete bin '{b.name}' — {unit_count} active unit(s) assigned.", "danger")
        return redirect(url_for('stock_bp.manage_bins', location_id=location_id))
    db.session.delete(b)
    db.session.commit()
    flash(f"Bin '{b.name}' deleted.", "success")
    return redirect(url_for('stock_bp.manage_bins', location_id=location_id))


@stock_bp.route('/location/<int:location_id>/bins/<int:bin_id>/qr')
@login_required
def bin_qr_label(location_id, bin_id):
    b = Bin.query.filter_by(id=bin_id, location_id=location_id, tenant_id=current_user.tenant_id).first_or_404()
    # Generate QR code for the bin_contents URL
    bin_url = url_for('stock_bp.bin_contents', bin_code=b.name, bin_id=b.id, _external=True)
    qr = qrcode.QRCode(version=1, box_size=6, border=4)
    qr.add_data(bin_url)
    qr.make(fit=True)
    img = qr.make_image(fill_color='black', back_color='white')
    buf = BytesIO()
    img.save(buf, format='PNG')
    qr_b64 = base64.b64encode(buf.getvalue()).decode()
    return render_template('bin_qr_label.html', bin=b, location=b.location, qr_b64=qr_b64, bin_url=bin_url)


# --- Location Contents (bins within a location) ---
@stock_bp.route('/location/<int:location_id>/contents')
@login_required
def location_contents(location_id):
    location = Location.query.filter_by(id=location_id, tenant_id=current_user.tenant_id).first_or_404()

    # Managed bins (Bin model) with unit counts via bin_id
    managed_bins = Bin.query.filter_by(location_id=location_id, tenant_id=current_user.tenant_id).order_by(Bin.name).all()
    bin_ids = [b.id for b in managed_bins]

    # Count units per bin_id
    bin_stats = {}
    if bin_ids:
        rows = db.session.query(
            ProductInstance.bin_id,
            func.count(ProductInstance.id).label('total'),
            func.sum(db.case((ProductInstance.status == 'processed', 1), else_=0)).label('processed'),
            func.sum(db.case((ProductInstance.status == 'unprocessed', 1), else_=0)).label('unprocessed'),
            func.sum(db.case((ProductInstance.status == 'under_process', 1), else_=0)).label('under_process'),
            func.sum(db.case((ProductInstance.status.in_(['idle', 'disputed']), 1), else_=0)).label('other'),
        ).join(Product).filter(
            ProductInstance.bin_id.in_(bin_ids),
            ProductInstance.is_sold == False,
            Product.tenant_id == current_user.tenant_id
        ).group_by(ProductInstance.bin_id).all()
        for row in rows:
            bin_stats[row.bin_id] = row

    # Legacy: shelf_bin-only instances (bin_id is NULL) still in this location
    legacy_rows = db.session.query(
        ProductInstance.shelf_bin,
        func.count(ProductInstance.id).label('total'),
        func.sum(db.case((ProductInstance.status == 'processed', 1), else_=0)).label('processed'),
        func.sum(db.case((ProductInstance.status == 'unprocessed', 1), else_=0)).label('unprocessed'),
        func.sum(db.case((ProductInstance.status == 'under_process', 1), else_=0)).label('under_process'),
        func.sum(db.case((ProductInstance.status.in_(['idle', 'disputed']), 1), else_=0)).label('other'),
    ).join(Product).filter(
        ProductInstance.location_id == location_id,
        ProductInstance.bin_id == None,
        ProductInstance.is_sold == False,
        Product.tenant_id == current_user.tenant_id
    ).group_by(ProductInstance.shelf_bin).order_by(ProductInstance.shelf_bin).all()

    # Recent movement log
    movements = ProductProcessLog.query.join(
        ProductInstance, ProductProcessLog.product_instance_id == ProductInstance.id
    ).join(Product, ProductInstance.product_id == Product.id).filter(
        ProductProcessLog.action.in_(['location_move', 'scan_move']),
        ProductInstance.location_id == location_id,
        Product.tenant_id == current_user.tenant_id
    ).order_by(ProductProcessLog.moved_at.desc()).limit(50).all()

    return render_template(
        'location_contents.html',
        location=location,
        managed_bins=managed_bins,
        bin_stats=bin_stats,
        legacy_rows=legacy_rows,
        movements=movements,
    )


# --- Bin Autocomplete (uses Bin model, falls back to shelf_bin strings) ---
@stock_bp.route('/bins/autocomplete')
@login_required
def bins_autocomplete():
    location_id = request.args.get('location_id', type=int)
    q = request.args.get('q', '').strip().upper()
    base = Bin.query.filter_by(tenant_id=current_user.tenant_id)
    if location_id:
        base = base.filter_by(location_id=location_id)
    if q:
        base = base.filter(Bin.name.ilike(f'{q}%'))
    bins = [{'id': b.id, 'name': b.name, 'location_id': b.location_id} for b in base.order_by(Bin.name).limit(20).all()]
    return jsonify(bins)


# --- Bins for a location (JSON, used by dynamic dropdowns) ---
@stock_bp.route('/bins/for_location')
@login_required
def bins_for_location():
    location_id = request.args.get('location_id', type=int)
    if not location_id:
        return jsonify([])
    bin_type = request.args.get('bin_type', '').strip()
    q = Bin.query.filter_by(location_id=location_id, tenant_id=current_user.tenant_id)
    if bin_type in ('units', 'parts'):
        q = q.filter_by(bin_type=bin_type)
    bins = q.order_by(Bin.name).all()
    return jsonify([{'id': b.id, 'name': b.name} for b in bins])


# --- AJAX: create a bin on-the-fly from a dropdown ---
@stock_bp.route('/bins/create', methods=['POST'])
@login_required
def create_bin_ajax():
    data = request.get_json() or {}
    name = (data.get('name') or '').strip().upper()
    location_id = data.get('location_id')
    if location_id:
        location_id = int(location_id)
    if not name or not location_id:
        return jsonify({'error': 'name and location_id are required'}), 400
    location = Location.query.filter_by(id=location_id, tenant_id=current_user.tenant_id).first()
    if not location:
        return jsonify({'error': 'Location not found'}), 404
    existing = Bin.query.filter_by(
        name=name, location_id=location_id, tenant_id=current_user.tenant_id
    ).first()
    if existing:
        return jsonify({'id': existing.id, 'name': existing.name, 'existed': True})
    b = Bin(name=name, location_id=location_id, tenant_id=current_user.tenant_id,
            created_at=get_now_for_tenant())
    db.session.add(b)
    db.session.commit()
    return jsonify({'id': b.id, 'name': b.name, 'existed': False})


# --- All Bins overview ---
@stock_bp.route('/bins')
@login_required
def all_bins():
    from collections import defaultdict
    locations = Location.query.filter_by(
        tenant_id=current_user.tenant_id
    ).order_by(Location.name).all()
    bins = Bin.query.filter_by(
        tenant_id=current_user.tenant_id
    ).order_by(Bin.location_id, Bin.name).all()
    bin_ids = [b.id for b in bins]
    counts = {}
    status_counts = {}
    if bin_ids:
        for row in db.session.query(
            ProductInstance.bin_id,
            func.count(ProductInstance.id).label('total'),
            func.sum(db.case((ProductInstance.status == 'processed', 1), else_=0)).label('processed'),
            func.sum(db.case((ProductInstance.status == 'unprocessed', 1), else_=0)).label('unprocessed'),
            func.sum(db.case((ProductInstance.status == 'under_process', 1), else_=0)).label('under_process'),
        ).join(Product).filter(
            ProductInstance.bin_id.in_(bin_ids),
            ProductInstance.is_sold == False,
            Product.tenant_id == current_user.tenant_id
        ).group_by(ProductInstance.bin_id).all():
            counts[row.bin_id] = row.total
            status_counts[row.bin_id] = {
                'processed': row.processed or 0,
                'unprocessed': row.unprocessed or 0,
                'under_process': row.under_process or 0,
            }
    bins_by_loc = defaultdict(list)
    for b in bins:
        bins_by_loc[b.location_id].append(b)
    return render_template(
        'all_bins.html',
        locations=locations,
        bins_by_loc=bins_by_loc,
        counts=counts,
        status_counts=status_counts,
    )


# --- Bin Detail: proper /stock/bin/<id> page ---
@stock_bp.route('/bin/<int:bin_id>')
@login_required
def bin_detail(bin_id):
    b = Bin.query.filter_by(id=bin_id, tenant_id=current_user.tenant_id).first_or_404()
    status_filter = request.args.get('status', '')
    query = ProductInstance.query.join(Product).filter(
        ProductInstance.bin_id == bin_id,
        ProductInstance.is_sold == False,
        Product.tenant_id == current_user.tenant_id,
    )
    if status_filter:
        query = query.filter(ProductInstance.status == status_filter)
    instances = query.options(
        joinedload(ProductInstance.product),
        joinedload(ProductInstance.assigned_user),
    ).order_by(ProductInstance.created_at.desc()).all()
    # Bins in the same location for quick-move target
    same_loc_bins = Bin.query.filter(
        Bin.location_id == b.location_id,
        Bin.tenant_id == current_user.tenant_id,
        Bin.id != bin_id,
    ).order_by(Bin.name).all()
    # All locations + bins for cross-location move
    locations = Location.query.filter_by(tenant_id=current_user.tenant_id).order_by(Location.name).all()
    # Parts stored in this bin
    from inventory_flask_app.models import PartStock, Part
    part_stocks = (
        PartStock.query
        .join(Part, PartStock.part_id == Part.id)
        .filter(
            PartStock.bin_id == bin_id,
            PartStock.quantity > 0,
            Part.tenant_id == current_user.tenant_id,
        )
        .options(db.joinedload(PartStock.part))
        .order_by(Part.name)
        .all()
    )
    return render_template(
        'bin_detail.html',
        bin=b,
        instances=instances,
        status_filter=status_filter,
        same_loc_bins=same_loc_bins,
        locations=locations,
        part_stocks=part_stocks,
    )


# --- Bin Detail: AJAX move single unit ---
@stock_bp.route('/bin/<int:bin_id>/move_unit', methods=['POST'])
@login_required
def bin_move_unit(bin_id):
    Bin.query.filter_by(id=bin_id, tenant_id=current_user.tenant_id).first_or_404()
    data = request.get_json() or {}
    instance_id = data.get('instance_id')
    target_bin_id = data.get('target_bin_id')
    instance = ProductInstance.query.join(Product).filter(
        ProductInstance.id == instance_id,
        Product.tenant_id == current_user.tenant_id,
    ).first()
    if not instance:
        return jsonify({'error': 'Instance not found'}), 404
    if target_bin_id:
        target_bin = Bin.query.filter_by(id=int(target_bin_id), tenant_id=current_user.tenant_id).first()
        if not target_bin:
            return jsonify({'error': 'Target bin not found'}), 404
        instance.bin_id = target_bin.id
        instance.shelf_bin = target_bin.name
        instance.location_id = target_bin.location_id
    else:
        instance.bin_id = None
        instance.shelf_bin = None
    db.session.commit()
    return jsonify({'success': True, 'bin_name': target_bin.name if target_bin_id else None})


# --- Bin Detail: bulk move ---
@stock_bp.route('/bin/<int:bin_id>/bulk_move', methods=['POST'])
@login_required
def bin_bulk_move(bin_id):
    b = Bin.query.filter_by(id=bin_id, tenant_id=current_user.tenant_id).first_or_404()
    instance_ids = request.form.getlist('instance_ids')
    instance_ids = [int(i) for i in instance_ids if i]
    target_bin_id = request.form.get('target_bin_id', type=int)
    if not instance_ids:
        flash("No units selected.", "warning")
        return redirect(url_for('stock_bp.bin_detail', bin_id=bin_id))
    target_bin = None
    if target_bin_id:
        target_bin = Bin.query.filter_by(id=target_bin_id, tenant_id=current_user.tenant_id).first()
    updated = 0
    for inst in ProductInstance.query.join(Product).filter(
        ProductInstance.id.in_(instance_ids),
        Product.tenant_id == current_user.tenant_id,
    ).all():
        if target_bin:
            inst.bin_id = target_bin.id
            inst.shelf_bin = target_bin.name
            inst.location_id = target_bin.location_id
        else:
            inst.bin_id = None
            inst.shelf_bin = None
        updated += 1
    db.session.commit()
    dest = target_bin.name if target_bin else 'no bin'
    flash(f"Moved {updated} unit(s) to {dest}.", "success")
    return redirect(url_for('stock_bp.bin_detail', bin_id=bin_id))


# --- Bin Detail: CSV export ---
@stock_bp.route('/bin/<int:bin_id>/export')
@login_required
def bin_export_csv(bin_id):
    b = Bin.query.filter_by(id=bin_id, tenant_id=current_user.tenant_id).first_or_404()
    instances = ProductInstance.query.join(Product).filter(
        ProductInstance.bin_id == bin_id,
        ProductInstance.is_sold == False,
        Product.tenant_id == current_user.tenant_id,
    ).options(joinedload(ProductInstance.product), joinedload(ProductInstance.assigned_user)).all()
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(['Serial', 'Asset', 'Item Name', 'Make', 'Model', 'CPU', 'RAM',
                     'Status', 'Stage', 'Team', 'Assigned To', 'Bin', 'Location'])
    for i in instances:
        writer.writerow([
            i.serial, i.asset or '',
            i.product.item_name if i.product else '',
            i.product.make if i.product else '',
            i.product.model if i.product else '',
            i.product.cpu if i.product else '',
            i.product.ram if i.product else '',
            i.status, i.process_stage or '', i.team_assigned or '',
            i.assigned_user.username if i.assigned_user else '',
            b.name, b.location.name if b.location else '',
        ])
    resp = make_response(output.getvalue())
    resp.headers['Content-Type'] = 'text/csv'
    resp.headers['Content-Disposition'] = f'attachment; filename=bin_{b.name}.csv'
    return resp


# --- Export grouped summary as Excel ---
@stock_bp.route('/export_grouped_summary')
@login_required
def export_grouped_summary():
    from openpyxl import Workbook
    from io import BytesIO
    from flask import send_file
    from sqlalchemy import or_

    # Reuse the filters used in under_process route
    status = request.args.get('status')
    model = request.args.get('model')
    processor = request.args.get('processor')
    serial_search = request.args.get('serial_search', '').strip()
    stage_filter = request.args.get('stage')
    team_filter = request.args.get('team')
    location_id = request.args.get('location_id')
    bin_search = request.args.get('bin_search', '').strip()
    ram_filter = request.args.get('ram')
    disk_filter = request.args.get('disk1size')

    if not status or status == 'all':
        query = ProductInstance.query
    else:
        if status == 'unprocessed':
            query = ProductInstance.query.filter_by(status='unprocessed', is_sold=False)
        elif status == 'under_process':
            query = ProductInstance.query.filter_by(status='under_process', is_sold=False)
        elif status == 'processed':
            query = ProductInstance.query.filter_by(status='processed', is_sold=False)
        else:
            query = ProductInstance.query.filter(ProductInstance.status == status)

    # Join and apply filters
    query = query.join(Product)
    if model:
        query = query.filter(Product.model.ilike(f"%{model}%"))
    if processor:
        query = query.filter(Product.cpu == processor)
    if ram_filter:
        query = query.filter(Product.ram == ram_filter)
    if disk_filter:
        query = query.filter(Product.disk1size == disk_filter)
    if serial_search:
        query = query.filter(or_(
            ProductInstance.serial.ilike(f"%{serial_search}%"),
            ProductInstance.asset.ilike(f"%{serial_search}%")
        ))
    if stage_filter:
        query = query.filter(ProductInstance.process_stage == stage_filter)
    if team_filter:
        query = query.filter(ProductInstance.team_assigned.ilike(f"%{team_filter}%"))
    if location_id:
        query = query.filter(ProductInstance.location_id == int(location_id))
    if bin_search:
        query = query.filter(ProductInstance.shelf_bin.ilike(f"%{bin_search}%"))

    query = query.filter(Product.tenant_id == current_user.tenant_id)

    all_instances = query.options(
        joinedload(ProductInstance.product)
    ).all()

    # Group by (model, cpu)
    from collections import defaultdict
    grouped = defaultdict(list)
    for instance in all_instances:
        if instance.product:
            key = (instance.product.model, instance.product.cpu)
            grouped[key].append(instance)

    # Prepare Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Grouped Summary"
    ws.append(["Model", "CPU", "Count"])

    for (model, cpu), instances in grouped.items():
        ws.append([model, cpu, len(instances)])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        download_name="grouped_inventory_summary.xlsx",
        as_attachment=True
    )
@stock_bp.route('/group_view')
@login_required
def group_view_page():
    model = request.args.get("model", "").strip()
    cpu = request.args.get("cpu", "").strip()
    if not model or not cpu:
        return redirect(url_for('stock_bp.under_process'))

    # --- Enhanced filtering logic for group_view ---
    ram = request.args.get("ram", "").strip()
    gpu = request.args.get("gpu", "").strip()
    display = request.args.get("display", "").strip()
    stage = request.args.get("stage", "").strip()
    status = request.args.get("status", "").strip()
    location = request.args.get("location", "").strip()

    # Use outerjoin(Location) so units with no location are included
    instances = ProductInstance.query.join(Product).outerjoin(Location).filter(
        Product.model == model,
        Product.cpu == cpu,
        Product.tenant_id == current_user.tenant_id
    )

    # Support filtering for aged units when filter=aged is passed in the query string
    filter_mode = request.args.get("filter")
    if filter_mode == "aged":
        threshold = datetime.utcnow() - timedelta(days=60)
        instances = instances.filter(ProductInstance.created_at <= threshold)

    if ram:
        instances = instances.filter(Product.ram == ram)
    if gpu:
        instances = instances.filter(Product.gpu1 == gpu)
    if display:
        instances = instances.filter(Product.display == display)
    if stage:
        instances = instances.filter(ProductInstance.process_stage == stage)
    if status and status != "all":
        instances = instances.filter(ProductInstance.status == status)
    if location:
        instances = instances.filter(Location.name == location)
    instances = instances.options(
        joinedload(ProductInstance.product).joinedload(Product.vendor),
        joinedload(ProductInstance.location),
        joinedload(ProductInstance.po),
    ).all()

    # Sort units
    unit_sort     = request.args.get('sort', 'serial')
    unit_sort_dir = request.args.get('sort_dir', 'asc')
    unit_reverse  = (unit_sort_dir == 'desc')
    _now_utc = datetime.utcnow()
    _sort_key = {
        'serial':        lambda i: (i.serial or '').lower(),
        'status':        lambda i: (i.status or '').lower(),
        'process_stage': lambda i: (i.process_stage or '').lower(),
        'location':      lambda i: (i.location.name if i.location else '').lower(),
        'grade':         lambda i: (i.product.grade if i.product else '').lower(),
        'age_days':      lambda i: ((_now_utc - i.created_at).days if i.created_at else 0),
        'asking_price':  lambda i: (i.asking_price or 0),
    }.get(unit_sort, lambda i: (i.serial or '').lower())
    instances = sorted(instances, key=_sort_key, reverse=unit_reverse)

    from inventory_flask_app.models import TenantSettings
    tenant_settings = TenantSettings.query.filter_by(tenant_id=current_user.tenant_id).all()
    settings_dict = {s.key: s.value for s in tenant_settings}
    column_order = settings_dict.get("column_order_instance_table")

    default_columns = [
        "asset", "serial", "Item Name", "make", "model", "display", "cpu", "ram",
        "gpu1", "gpu2", "Grade", "location", "status", "process_stage",
        "team_assigned", "shelf_bin", "is_sold", "age_days", "asking_price",
        "vendor", "po_number", "label", "action"
    ]
    for _nc in ('age_days', 'asking_price', 'vendor', 'po_number'):
        settings_dict.setdefault(f'show_column_{_nc}', 'false')

    # Paginate
    _per_page = 50
    _total_count = len(instances)
    _total_pages = max(1, (_total_count + _per_page - 1) // _per_page)
    _page = min(max(request.args.get('page', 1, type=int), 1), _total_pages)
    instances_page = instances[(_page - 1) * _per_page : _page * _per_page]

    return render_template(
        "group_view.html",
        instances=instances_page,
        settings=settings_dict,
        location_name=location,
        model=model,
        cpu=cpu,
        unit_sort=unit_sort,
        unit_sort_dir=unit_sort_dir,
        now_utc=datetime.utcnow(),
        page=_page,
        total_pages=_total_pages,
        total_count=_total_count,
        per_page=_per_page,
    )
# --- Aged Inventory Route ---
@stock_bp.route('/inventory/aged')
@login_required
def aged_inventory():
    from collections import defaultdict
    from inventory_flask_app.models import TenantSettings

    threshold_days = 60
    cutoff_date = datetime.utcnow() - timedelta(days=threshold_days)

    instances = ProductInstance.query.join(Product).options(
        joinedload(ProductInstance.product),
        joinedload(ProductInstance.location)
    ).filter(
        Product.tenant_id == current_user.tenant_id,
        ProductInstance.created_at <= cutoff_date,
        ProductInstance.is_sold == False
    ).all()

    grouped = defaultdict(list)
    for i in instances:
        key = (i.product.model if i.product else "", i.product.cpu if i.product else "")
        grouped[key].append(i)

    grouped_instances = []
    for key, group in grouped.items():
        product_id = next((i.product.id for i in group if i.product and i.product.id), None)
        grouped_instances.append({
            "model": key[0],
            "cpu": key[1],
            "count": len(group),
            "instances": group,
            "product_id": product_id
        })

    tenant_settings = TenantSettings.query.filter_by(tenant_id=current_user.tenant_id).all()
    settings_dict = {s.key: s.value for s in tenant_settings}
    column_order = settings_dict.get("column_order_instance_table")

    default_columns = [
        "asset", "serial", "Item Name", "make", "model", "display", "cpu", "ram",
        "gpu1", "gpu2", "Grade", "location", "status", "process_stage",
        "team_assigned", "shelf_bin", "is_sold", "label", "action"
    ]

    _alias_map = {
        # Old API names
        "model_number": "model",
        "product": "Item Name",
        "processor": "cpu",
        "video_card": "gpu1",
        "resolution": "display",
        "screen_size": "display",
        # Case/format variants saved by admin column reorder
        "item_name": "Item Name",
        "item name": "Item Name",
        "grade": "Grade",
        "team": "team_assigned",
        "gpu": "gpu1",
    }
    if column_order:
        seen = set()
        existing = []
        for col in column_order.split(","):
            col = _alias_map.get(col.strip(), col.strip())
            if col and col not in seen:
                seen.add(col)
                existing.append(col)
        for col in default_columns:
            if col not in seen:
                seen.add(col)
                existing.append(col)
        column_order = ",".join(existing)
    else:
        column_order = ",".join(default_columns)

    settings_dict["column_order_instance_table"] = column_order

    return render_template(
        "aged_inventory.html",
        grouped_instances=grouped_instances,
        settings=settings_dict,
        threshold_days=threshold_days
    )